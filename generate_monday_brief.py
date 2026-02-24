#!/usr/bin/env python3
"""
ExecSignals — The Monday Brief Generator.

Produces the full weekly deliverable for executive recruiters:
  1. Excel workbook (Top Leads + Market Intel)
  2. Market Intel PDF one-pager (HTML for browser → Print → Save as PDF)
  3. Email (HTML + plain text)
  4. CSV (raw leads)

Usage:
    python3 generate_monday_brief.py --preview
    python3 generate_monday_brief.py --preview --top 20
    python3 generate_monday_brief.py --send --resend-key KEY

Dependencies:
    pip install openpyxl  (Excel generation)
    pip install resend    (only for --send mode)
"""

import argparse
import html
import math
import os
import sqlite3
import sys
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import scoring + lead fetching from existing pipeline
from generate_hot_leads import (
    fetch_hot_leads,
    score_lead,
    format_salary,
    format_location,
    extract_hiring_signal,
    extract_team_structure,
    extract_extra_signals,
    extract_key_tools,
    generate_csv,
)

# ─── Configuration ────────────────────────────────────────────────────────────

DEFAULT_DB = "/Users/rome/Documents/projects/scrapers/master/data/jobs.db"
VP_TIERS = ("vp", "svp", "evp", "c_level")

# ─── Display Name Mappings ────────────────────────────────────────────────────

FUNCTION_TO_ROLE = {
    "sales": "VP Sales",
    "finance": "CFO",
    "engineering": "VP Engineering",
    "marketing": "VP Marketing",
    "operations": "VP Operations",
    "product": "VP Product",
    "people": "VP People/HR",
    "data": "VP Data",
    "legal": "VP Legal",
}

ROLE_ORDER = [
    "VP Sales", "CFO", "VP Engineering", "VP Marketing",
    "VP Operations", "VP Product", "VP People/HR",
]

INDUSTRY_MAP = {
    "Banks And Financial Services": "Financial Services",
    "Health Care": "Healthcare",
    "Education And Schools": "Education",
    "Internet And Software": "Software / SaaS",
    "Government": "Government",
    "Organization": "Nonprofit",
    "Consulting And Business Services": "Consulting",
    "Media News And Publishing": "Media",
    "Restaurants Travel And Leisure": "Hospitality",
    "Insurance": "Insurance",
    "Real Estate": "Real Estate",
    "Transport And Freight": "Transport",
    "Retail": "Retail",
    "Consumer Goods And Services": "Consumer Goods",
    "Industrial Manufacturing": "Manufacturing",
    "Energy Mining And Utilities": "Energy",
    "Telecommunications": "Telecom",
    "Construction": "Construction",
    "Agriculture": "Agriculture",
    "Automotive": "Automotive",
}

STAGE_MAP = {
    "enterprise": "Enterprise / Public",
    "public": "Enterprise / Public",
    "late_stage": "Late Stage",
    "late stage": "Late Stage",
    "growth": "Growth",
    "series_a": "Growth",
    "series_b": "Growth",
    "series_c": "Late Stage",
    "series_d": "Late Stage",
    "early_stage": "Early Stage",
    "early stage": "Early Stage",
    "seed": "Early Stage",
    "startup": "Early Stage",
}

SENIORITY_DISPLAY = {
    "c_level": "C-Level",
    "evp": "EVP",
    "svp": "SVP",
    "vp": "VP",
}

# Roles where "Reports CRO" signal makes sense
SALES_FUNCTIONS = {"sales", "business_development", "revenue", "partnerships"}
SALES_TITLE_KEYWORDS = {"sales", "revenue", "business development", "account",
                         "partnerships", "growth", "commercial", "sdr", "bdr"}


COMPANY_NAME_OVERRIDES = {
    "jpmorganchase": "JPMorgan Chase",
    "jpmorgan chase": "JPMorgan Chase",
    "blackrock": "BlackRock",
    "mckinsey": "McKinsey",
    "deloitte": "Deloitte",
    "pwc": "PwC",
    "ey": "EY",
    "kpmg": "KPMG",
    "ibm": "IBM",
    "att": "AT&T",
    "cvs": "CVS Health",
    "ge": "GE",
    "hp": "HP",
    "sap": "SAP",
    "bny": "BNY Mellon",
    "ymca": "YMCA",
    "usaa": "USAA",
    "hca": "HCA Healthcare",
    "bcg": "BCG",
}

# Companies to exclude from Top Hiring Companies (venture studios, job boards, etc.)
COMPANY_BLOCKLIST = {
    "futuresight",  # Venture studio posting co-founder roles, not an employer
}


def format_company_name(name):
    """Title-case a company name, handling special cases."""
    if not name:
        return "Confidential"
    # Check overrides first
    lower = name.lower().strip()
    if lower in COMPANY_NAME_OVERRIDES:
        return COMPANY_NAME_OVERRIDES[lower]
    # Already looks properly cased (has mixed case, not ALL CAPS)
    has_upper = any(c.isupper() for c in name[1:])
    has_lower = any(c.islower() for c in name[1:])
    if has_upper and has_lower:
        return name
    # Title-case with common acronym handling
    words = name.title().split()
    acronyms = {"Llc", "Inc", "Llp", "Lp", "Pc", "Pllc", "Dds", "Md",
                "Nyc", "Usa", "Us", "Ai"}
    result = []
    for w in words:
        if w.rstrip(".,") in acronyms:
            result.append(w.upper())
        else:
            result.append(w)
    return " ".join(result)


def filter_signals_for_role(lead):
    """Filter out 'reports_cro' signal for non-sales roles."""
    title_lower = (lead.get("title") or "").lower()
    func = (lead.get("function_category") or "").lower()
    is_sales_role = (func in SALES_FUNCTIONS or
                     any(kw in title_lower for kw in SALES_TITLE_KEYWORDS))
    if is_sales_role:
        return lead.get("signals", [])
    return [s for s in lead.get("signals", []) if s["signal_id"] != "reports_cro"]


def estimate_placement_fee(lead, pct=0.25):
    """Estimate recruiter placement fee (25% of salary midpoint)."""
    sal_min = lead.get("annual_salary_min") or 0
    sal_max = lead.get("annual_salary_max") or 0
    if sal_max > 0:
        midpoint = (sal_min + sal_max) / 2 if sal_min > 0 else sal_max
    elif sal_min > 0:
        midpoint = sal_min
    else:
        return None
    fee = midpoint * pct
    if fee >= 1000:
        return f"${fee / 1000:.0f}K"
    return f"${fee:,.0f}"


# Known executive search / staffing firms — their postings are retained searches
SEARCH_FIRMS = {
    "korn ferry", "heidrick & struggles", "heidrick and struggles",
    "spencer stuart", "russell reynolds", "egon zehnder",
    "boyden", "odgers berndtson", "stanton chase", "dhr international",
    "jm search", "witt/kieffer", "wittkieffer", "diversified search",
    "caldwell partners", "isaacson miller",
    "robert half", "randstad", "adecco", "manpower", "manpowergroup",
    "kelly services", "hays", "page executive", "michael page",
}


def is_search_firm(lead):
    """Detect if the posting company is a known executive search firm."""
    name = (lead.get("company_name_normalized") or "").lower().strip()
    return name in SEARCH_FIRMS


def clean_location(lead):
    """Clean up messy location formatting from the scraper.

    Fixes: 'Remote, US' → 'Remote', 'Elgin, IL, US' → 'Elgin, IL',
    'Boston, (Remote)' → 'Boston (Remote)', removes redundant 'Remote'.
    """
    metro = lead.get("location_metro")
    raw = lead.get("location_raw") or ""
    is_remote = lead.get("is_remote")
    loc_type = (lead.get("location_type") or "").lower()

    # Clean the raw location: strip trailing ", US"
    if raw.endswith(", US"):
        raw = raw[:-4]
    # Strip leading "Remote, " if we'll add (Remote) anyway
    if raw.startswith("Remote, "):
        raw = raw[8:]
    if raw == "Remote":
        raw = ""
    if raw == "US":
        raw = ""

    # Build clean location
    parts = []
    if metro:
        parts.append(metro)
    elif raw:
        parts.append(raw)

    if is_remote or "remote" in loc_type:
        if parts:
            parts.append("(Remote)")
        else:
            parts.append("Remote")
    elif "hybrid" in loc_type:
        parts.append("(Hybrid)")

    return " ".join(parts) if parts else "Location not specified"


def correct_seniority(lead):
    """Override seniority_tier when the scraper misclassifies based on C-suite keyword proximity.

    The scraper tags 'Senior Director, CEO Initiatives' as c_level because it mentions CEO.
    This function checks if the title itself is actually a C-level role.
    """
    if lead.get("seniority_tier") != "c_level":
        return  # Only fix c_level misclassifications
    title = (lead.get("title") or "").lower().strip()
    # Patterns that confirm actual C-level role
    if title.startswith("chief "):
        return
    if title.startswith("ceo") or title.startswith("cfo") or title.startswith("coo"):
        return
    if title.startswith("cto") or title.startswith("cio") or title.startswith("cmo"):
        return
    if title.startswith("cro") or title.startswith("cso") or title.startswith("cpo"):
        return
    if title.startswith("president") or "president and" in title or "president &" in title:
        return
    if title.startswith("executive director"):
        return
    if title.startswith("general counsel") or title.startswith("managing director"):
        return
    if "founding" in title and ("president" in title or "ceo" in title or "chief" in title):
        return
    # Titles that mention C-suite but aren't C-level roles
    downgrade_patterns = [
        "director,", "director ", "sr. director", "senior director",
        "manager,", "manager ", "sr. manager", "senior manager",
        "coordinator", "analyst", "specialist", "associate",
        "advisor", "intern", "assistant",
    ]
    for pattern in downgrade_patterns:
        if title.startswith(pattern) or f" {pattern}" in title[:30]:
            lead["seniority_tier"] = "vp"
            return
    # If title doesn't start with a C-suite keyword and doesn't match
    # any known pattern, keep the scraper's classification


def is_false_positive(lead):
    """Filter out leads that aren't real job openings (training programs, internships, etc.)."""
    title = (lead.get("title") or "").lower()
    # Certification/training programs masquerading as exec roles
    if "certification program" in title:
        return True
    if "future leaders program" in title:
        return True
    # Internships and trainee roles
    if "internship" in title:
        return True
    if title.startswith("intern ") or title.startswith("intern-"):
        return True
    if "in-training" in title or "-in-training" in title:
        return True
    return False


def apply_freshness_bonus(lead, ref_date):
    """Add freshness bonus to score. 0-2 days = +10, 3-4 days = +5, 5-7 = +2."""
    date_posted = lead.get("date_posted")
    if not date_posted:
        return
    try:
        posted_dt = datetime.strptime(str(date_posted)[:10], "%Y-%m-%d")
        days_ago = (ref_date - posted_dt).days
    except ValueError:
        return
    if days_ago <= 2:
        lead["score"] += 10
    elif days_ago <= 4:
        lead["score"] += 5
    elif days_ago <= 7:
        lead["score"] += 2
    # Penalize stale leads (>14 days)
    elif days_ago > 14:
        lead["score"] -= 5


# ─── Excel Brand Palette ─────────────────────────────────────────────────────

NAVY = "0C0F1A"
AMBER = "D4A054"
LIGHT_AMBER = "E2A84B"
BLUE = "5B8DEF"
DARK_TEXT = "1A1A1A"
WHITE = "FFFFFF"
LIGHT_BG = "F8F9FB"
BORDER_COLOR = "D0D5DD"
RED_FONT = "C0392B"
AMBER_FONT = "B7791F"
GRAY_FONT = "888888"
GREEN_FONT = "27763D"
SECTION_BG = "1A1F2E"

XL_HEADER_FONT = Font(name="Plus Jakarta Sans", bold=True, color=WHITE, size=11)
XL_BODY_FONT = Font(name="Plus Jakarta Sans", color=DARK_TEXT, size=10)
XL_BODY_BOLD = Font(name="Plus Jakarta Sans", color=DARK_TEXT, size=10, bold=True)
XL_LINK_FONT = Font(name="Plus Jakarta Sans", color="1155CC", size=10, underline="single")
XL_SECTION_FONT = Font(name="DM Serif Display", bold=True, color=AMBER, size=13)

XL_HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
XL_SCORE_GOLD = PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid")
XL_SCORE_BLUE = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
XL_SCORE_LIGHT = PatternFill(start_color=LIGHT_AMBER, end_color=LIGHT_AMBER, fill_type="solid")
XL_SCORE_GRAY = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
XL_ALT_ROW = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
XL_SECTION_FILL = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")

XL_THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

XL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
XL_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
XL_LEFT = Alignment(horizontal="left", vertical="center")
XL_RIGHT = Alignment(horizontal="right", vertical="center")


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 1: MARKET ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════════


def _percentile(values, pct):
    """Compute percentile from a sorted list of values."""
    if not values:
        return 0
    k = (len(values) - 1) * (pct / 100)
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return values[int(k)]
    return values[f] * (c - k) + values[c] * (k - f)


def compute_salary_benchmarks(conn, days=14):
    """Compute P25/Median/P75 salary benchmarks by role function.

    Uses ALL active jobs for current percentiles (broader sample),
    then compares recent vs prior 30-day windows for trend.
    """
    tier_placeholders = ",".join("?" for _ in VP_TIERS)
    # For trend: compare last 30 days vs 30-60 days ago (relative to latest data)
    ref_date = _get_data_reference_date(conn)
    trend_cutoff = (ref_date - timedelta(days=30)).strftime("%Y-%m-%d")
    trend_prior = (ref_date - timedelta(days=60)).strftime("%Y-%m-%d")

    benchmarks = []
    for func, role_name in FUNCTION_TO_ROLE.items():
        if role_name not in ROLE_ORDER:
            continue

        # Current: all active jobs with salary (no date filter for stable percentiles)
        rows = conn.execute(f"""
            SELECT annual_salary_max FROM jobs
            WHERE is_active = 1
              AND seniority_tier IN ({tier_placeholders})
              AND function_category = ?
              AND annual_salary_max > 0
            ORDER BY annual_salary_max
        """, (*VP_TIERS, func)).fetchall()

        salaries = [r[0] for r in rows]
        if len(salaries) < 3:
            continue

        p25 = _percentile(salaries, 25)
        median = _percentile(salaries, 50)
        p75 = _percentile(salaries, 75)

        # Recent window for trend numerator
        recent_rows = conn.execute(f"""
            SELECT annual_salary_max FROM jobs
            WHERE is_active = 1
              AND seniority_tier IN ({tier_placeholders})
              AND function_category = ?
              AND annual_salary_max > 0
              AND date_posted >= ?
            ORDER BY annual_salary_max
        """, (*VP_TIERS, func, trend_cutoff)).fetchall()

        # Prior window for trend denominator
        prior_rows = conn.execute(f"""
            SELECT annual_salary_max FROM jobs
            WHERE is_active = 1
              AND seniority_tier IN ({tier_placeholders})
              AND function_category = ?
              AND annual_salary_max > 0
              AND date_posted >= ? AND date_posted < ?
            ORDER BY annual_salary_max
        """, (*VP_TIERS, func, trend_prior, trend_cutoff)).fetchall()

        recent_salaries = [r[0] for r in recent_rows]
        prior_salaries = [r[0] for r in prior_rows]

        if len(recent_salaries) >= 3 and len(prior_salaries) >= 3:
            recent_median = _percentile(recent_salaries, 50)
            prior_median = _percentile(prior_salaries, 50)
            trend_pct = round((recent_median - prior_median) / prior_median * 100, 1) if prior_median else 0
        else:
            trend_pct = 0

        benchmarks.append({
            "role": role_name,
            "p25": f"${int(p25 / 1000)}K",
            "median": f"${int(median / 1000)}K",
            "p75": f"${int(p75 / 1000)}K",
            "p25_raw": p25,
            "median_raw": median,
            "p75_raw": p75,
            "trend_pct": trend_pct,
            "trend_display": f"+{trend_pct}%" if trend_pct > 0 else f"{trend_pct}%",
            "count": len(salaries),
        })

    # Sort by ROLE_ORDER
    order_map = {r: i for i, r in enumerate(ROLE_ORDER)}
    benchmarks.sort(key=lambda x: order_map.get(x["role"], 99))
    return benchmarks


def _get_data_reference_date(conn):
    """Get the latest date_posted in the DB as the reference point for WoW."""
    row = conn.execute("""
        SELECT MAX(date_posted) FROM jobs WHERE is_active = 1
    """).fetchone()
    if row and row[0]:
        return datetime.strptime(str(row[0])[:10], "%Y-%m-%d")
    return datetime.now()


def compute_industry_velocity(conn, days=30):
    """Compute VP+ hiring velocity by industry with WoW change.

    Uses the full `days` window for total counts, but always compares
    last 7 days vs prior 7 days for a true week-over-week percentage.
    """
    tier_placeholders = ",".join("?" for _ in VP_TIERS)
    ref_date = _get_data_reference_date(conn)
    cutoff = (ref_date - timedelta(days=days)).strftime("%Y-%m-%d")

    # WoW: always 7-day windows regardless of `days`
    wow_current_start = (ref_date - timedelta(days=7)).strftime("%Y-%m-%d")
    wow_prior_start = (ref_date - timedelta(days=14)).strftime("%Y-%m-%d")
    ref_str = ref_date.strftime("%Y-%m-%d")

    # Total counts for the full window
    current = conn.execute(f"""
        SELECT company_industry, COUNT(*) as cnt FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND company_industry IS NOT NULL
          AND date_posted >= ?
        GROUP BY company_industry
        ORDER BY cnt DESC
    """, (*VP_TIERS, cutoff)).fetchall()

    # Last 7 days (for WoW numerator)
    wow_curr = conn.execute(f"""
        SELECT company_industry, COUNT(*) as cnt FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND company_industry IS NOT NULL
          AND date_posted >= ? AND date_posted <= ?
        GROUP BY company_industry
    """, (*VP_TIERS, wow_current_start, ref_str)).fetchall()

    # Prior 7 days (for WoW denominator)
    wow_prev = conn.execute(f"""
        SELECT company_industry, COUNT(*) as cnt FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND company_industry IS NOT NULL
          AND date_posted >= ? AND date_posted < ?
        GROUP BY company_industry
    """, (*VP_TIERS, wow_prior_start, wow_current_start)).fetchall()

    wow_curr_map = {r[0]: r[1] for r in wow_curr}
    wow_prev_map = {r[0]: r[1] for r in wow_prev}

    velocity = []
    for row in current:
        raw_name = row[0]
        display_name = INDUSTRY_MAP.get(raw_name, raw_name)
        count = row[1]
        wc = wow_curr_map.get(raw_name, 0)
        wp = wow_prev_map.get(raw_name, 0)
        wow = round((wc - wp) / wp * 100) if wp > 0 else 0

        velocity.append({
            "industry": display_name,
            "count": count,
            "wow_pct": wow,
            "wow_display": f"+{wow}%" if wow > 0 else f"{wow}%",
        })

    return velocity[:8]


def compute_top_companies(conn, days=30):
    """Compute top hiring companies with 'new this week' flags.

    Total counts use the full `days` window.
    'New' flag compares last 7 days vs prior 7 days.
    """
    tier_placeholders = ",".join("?" for _ in VP_TIERS)
    ref_date = _get_data_reference_date(conn)
    cutoff = (ref_date - timedelta(days=days)).strftime("%Y-%m-%d")

    # "New" detection: last 7 vs prior 7
    wow_current_start = (ref_date - timedelta(days=7)).strftime("%Y-%m-%d")
    wow_prior_start = (ref_date - timedelta(days=14)).strftime("%Y-%m-%d")
    ref_str = ref_date.strftime("%Y-%m-%d")

    # Count UNIQUE titles per company (not total posts) to avoid multi-location inflation
    current = conn.execute(f"""
        SELECT company_name_normalized,
               COUNT(DISTINCT title) as unique_roles
        FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND company_name_normalized IS NOT NULL
          AND date_posted >= ?
        GROUP BY company_name_normalized
        HAVING unique_roles >= 3
        ORDER BY unique_roles DESC
        LIMIT 30
    """, (*VP_TIERS, cutoff)).fetchall()

    # Companies present in prior 7-day window
    prior = conn.execute(f"""
        SELECT DISTINCT company_name_normalized FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND company_name_normalized IS NOT NULL
          AND date_posted >= ? AND date_posted < ?
    """, (*VP_TIERS, wow_prior_start, wow_current_start)).fetchall()

    prior_set = {r[0] for r in prior}

    companies = []
    for row in current:
        name = row[0]
        count = row[1]
        # Filter search firms from top companies list
        if name.lower().strip() in SEARCH_FIRMS:
            continue
        if name.lower().strip() in COMPANY_BLOCKLIST:
            continue
        is_new = name not in prior_set
        companies.append({
            "company": name,
            "count": count,
            "is_new": is_new,
        })

    return companies[:10]


def compute_geo_breakdown(conn, days=30):
    """Compute VP+ leads by metro area with WoW change.

    Total counts use the full `days` window.
    WoW compares last 7 days vs prior 7 days.
    """
    tier_placeholders = ",".join("?" for _ in VP_TIERS)
    ref_date = _get_data_reference_date(conn)
    cutoff = (ref_date - timedelta(days=days)).strftime("%Y-%m-%d")

    # WoW: always 7-day windows
    wow_current_start = (ref_date - timedelta(days=7)).strftime("%Y-%m-%d")
    wow_prior_start = (ref_date - timedelta(days=14)).strftime("%Y-%m-%d")
    ref_str = ref_date.strftime("%Y-%m-%d")

    # Total counts for full window — metro
    current_metro = conn.execute(f"""
        SELECT location_metro, COUNT(*) as cnt FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND location_metro IS NOT NULL
          AND (is_remote = 0 OR is_remote IS NULL)
          AND date_posted >= ?
        GROUP BY location_metro
        ORDER BY cnt DESC
        LIMIT 10
    """, (*VP_TIERS, cutoff)).fetchall()

    # Total remote count
    remote_count = conn.execute(f"""
        SELECT COUNT(*) FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND (is_remote = 1 OR location_type LIKE '%remote%')
          AND date_posted >= ?
    """, (*VP_TIERS, cutoff)).fetchone()[0]

    # WoW current 7 days — metro
    wow_curr_metro = conn.execute(f"""
        SELECT location_metro, COUNT(*) as cnt FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND location_metro IS NOT NULL
          AND (is_remote = 0 OR is_remote IS NULL)
          AND date_posted >= ? AND date_posted <= ?
        GROUP BY location_metro
    """, (*VP_TIERS, wow_current_start, ref_str)).fetchall()

    # WoW prior 7 days — metro
    wow_prev_metro = conn.execute(f"""
        SELECT location_metro, COUNT(*) as cnt FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND location_metro IS NOT NULL
          AND (is_remote = 0 OR is_remote IS NULL)
          AND date_posted >= ? AND date_posted < ?
        GROUP BY location_metro
    """, (*VP_TIERS, wow_prior_start, wow_current_start)).fetchall()

    # WoW remote
    wow_curr_remote = conn.execute(f"""
        SELECT COUNT(*) FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND (is_remote = 1 OR location_type LIKE '%remote%')
          AND date_posted >= ? AND date_posted <= ?
    """, (*VP_TIERS, wow_current_start, ref_str)).fetchone()[0]

    wow_prev_remote = conn.execute(f"""
        SELECT COUNT(*) FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND (is_remote = 1 OR location_type LIKE '%remote%')
          AND date_posted >= ? AND date_posted < ?
    """, (*VP_TIERS, wow_prior_start, wow_current_start)).fetchone()[0]

    wow_curr_map = {r[0]: r[1] for r in wow_curr_metro}
    wow_prev_map = {r[0]: r[1] for r in wow_prev_metro}

    geo = []
    for row in current_metro:
        metro = row[0]
        count = row[1]
        wc = wow_curr_map.get(metro, 0)
        wp = wow_prev_map.get(metro, 0)
        wow = round((wc - wp) / wp * 100) if wp > 0 else 0
        geo.append({
            "metro": metro,
            "count": count,
            "wow_pct": wow,
            "wow_display": f"+{wow}%" if wow > 0 else f"{wow}%",
        })

    # Add Remote bucket
    remote_wow = round((wow_curr_remote - wow_prev_remote) / wow_prev_remote * 100) if wow_prev_remote > 0 else 0
    geo.append({
        "metro": "Remote",
        "count": remote_count,
        "wow_pct": remote_wow,
        "wow_display": f"+{remote_wow}%" if remote_wow > 0 else f"{remote_wow}%",
    })

    # Sort by count descending
    geo.sort(key=lambda x: x["count"], reverse=True)
    return geo[:8]


def compute_company_stage(conn, days=7):
    """Compute company stage distribution for VP+ roles."""
    tier_placeholders = ",".join("?" for _ in VP_TIERS)
    ref_date = _get_data_reference_date(conn)
    cutoff = (ref_date - timedelta(days=days)).strftime("%Y-%m-%d")

    rows = conn.execute(f"""
        SELECT company_stage, COUNT(*) as cnt FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND date_posted >= ?
        GROUP BY company_stage
    """, (*VP_TIERS, cutoff)).fetchall()

    buckets = {
        "Enterprise / Public": 0,
        "Late Stage": 0,
        "Growth": 0,
        "Early Stage": 0,
        "Unknown": 0,
    }

    total = 0
    for row in rows:
        stage_raw = (row[0] or "").lower().strip()
        count = row[1]
        total += count

        mapped = STAGE_MAP.get(stage_raw)
        if mapped:
            buckets[mapped] += count
        elif stage_raw:
            # Try partial matching
            if "enterprise" in stage_raw or "public" in stage_raw:
                buckets["Enterprise / Public"] += count
            elif "late" in stage_raw or "series c" in stage_raw or "series d" in stage_raw:
                buckets["Late Stage"] += count
            elif "growth" in stage_raw or "series a" in stage_raw or "series b" in stage_raw:
                buckets["Growth"] += count
            elif "early" in stage_raw or "seed" in stage_raw or "startup" in stage_raw:
                buckets["Early Stage"] += count
            else:
                buckets["Unknown"] += count
        else:
            buckets["Unknown"] += count

    # Convert to percentages
    stages = []
    for name in ["Enterprise / Public", "Late Stage", "Growth", "Early Stage", "Unknown"]:
        pct = round(buckets[name] / total * 100) if total > 0 else 0
        stages.append({"stage": name, "count": buckets[name], "pct": pct})

    return stages


def compute_stack_trends(conn, days=7):
    """Compute top tools/stack mentioned in VP+ roles."""
    tier_placeholders = ",".join("?" for _ in VP_TIERS)
    ref_date = _get_data_reference_date(conn)
    cutoff = (ref_date - timedelta(days=days)).strftime("%Y-%m-%d")

    rows = conn.execute(f"""
        SELECT jt.tool_name, COUNT(DISTINCT jt.job_id) as cnt
        FROM job_tools jt
        JOIN jobs j ON jt.job_id = j.id
        WHERE j.is_active = 1
          AND j.seniority_tier IN ({tier_placeholders})
          AND j.date_posted >= ?
          AND jt.tool_name IS NOT NULL
          AND jt.tool_name <> ''
          AND LOWER(jt.tool_name) <> '_none'
        GROUP BY jt.tool_name
        ORDER BY cnt DESC
        LIMIT 8
    """, (*VP_TIERS, cutoff)).fetchall()

    # Get total VP+ roles this week for percentages
    total = conn.execute(f"""
        SELECT COUNT(*) FROM jobs
        WHERE is_active = 1
          AND seniority_tier IN ({tier_placeholders})
          AND date_posted >= ?
    """, (*VP_TIERS, cutoff)).fetchone()[0]

    tools = []
    for row in rows:
        pct = round(row[1] / total * 100) if total > 0 else 0
        tools.append({"tool": row[0], "count": row[1], "pct": pct})

    return tools


def compute_summary_stats(leads, analytics_geo):
    """Compute summary statistics from scored leads."""
    total = len(leads)
    if total == 0:
        return {"total": 0, "avg_salary": "$0K", "avg_score": 0, "growth_pct": 0,
                "seniority": {}, "segment": {}}

    # Average salary
    salaries = [l.get("annual_salary_max") or l.get("annual_salary_min") or 0 for l in leads]
    salaries = [s for s in salaries if s > 0]
    avg_salary = int(sum(salaries) / len(salaries) / 1000) if salaries else 0

    # Average score
    avg_score = int(sum(l["score"] for l in leads) / total)

    # Growth hire percentage
    growth_count = 0
    for lead in leads:
        for sig in lead.get("signals", []):
            if sig["signal_id"] == "growth_hire":
                growth_count += 1
                break
    growth_pct = round(growth_count / total * 100)

    # Seniority breakdown
    seniority = Counter()
    for lead in leads:
        tier = lead.get("seniority_tier", "unknown")
        display = SENIORITY_DISPLAY.get(tier, tier.replace("_", " ").title())
        seniority[display] += 1

    # Segment breakdown
    segment_display = {
        "enterprise": "Enterprise", "smb": "SMB", "mid_market": "Mid-Market",
        "fortune_500": "Fortune 500", "startup": "Startup",
    }
    segment = Counter()
    for lead in leads:
        for sig in lead.get("signals", []):
            if sig["signal_type"] == "segment":
                raw = sig["signal_id"]
                display = segment_display.get(raw, raw.replace("_", " ").title())
                segment[display] += 1
                break

    seg_total = sum(segment.values()) or 1
    segment_pct = {k: round(v / seg_total * 100) for k, v in segment.most_common(4)}

    # C-level count
    c_level_count = sum(1 for l in leads if l.get("seniority_tier") in ("c_level", "evp"))

    return {
        "total": total,
        "avg_salary": f"${avg_salary}K",
        "avg_salary_raw": avg_salary * 1000,
        "avg_score": avg_score,
        "growth_pct": growth_pct,
        "c_level_count": c_level_count,
        "seniority": dict(seniority.most_common()),
        "segment": segment_pct,
    }


def compute_remote_function_counts(conn, days=7):
    """Count remote VP+ roles by function_category for contextual stats."""
    tier_placeholders = ",".join("?" for _ in VP_TIERS)
    rows = conn.execute(f"""
        SELECT function_category, COUNT(*) as cnt
        FROM jobs
        WHERE seniority_tier IN ({tier_placeholders})
          AND is_active = 1
          AND (is_remote = 1 OR location_type LIKE '%remote%')
          AND date(date_scraped) >= date('now', '-{days} days')
          AND function_category IS NOT NULL
        GROUP BY function_category
    """, VP_TIERS).fetchall()
    return {r[0]: r[1] for r in rows}


def compute_all_analytics(conn, lead_days=30):
    """Compute all market analytics in one call.

    lead_days controls the lead selection window (passed from --days).
    Velocity/geo/companies always use 7-day windows for WoW comparison.
    """
    return {
        "salary_benchmarks": compute_salary_benchmarks(conn, days=14),
        "industry_velocity": compute_industry_velocity(conn, days=lead_days),
        "top_companies": compute_top_companies(conn, days=lead_days),
        "geo_breakdown": compute_geo_breakdown(conn, days=lead_days),
        "company_stage": compute_company_stage(conn, days=lead_days),
        "stack_trends": compute_stack_trends(conn, days=lead_days),
        "remote_function_counts": compute_remote_function_counts(conn, days=lead_days),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 6: SIGNAL NOTE GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════


def generate_signal_note(lead):
    """Generate a contextual signal note from lead data."""
    signals = {s["signal_id"] for s in lead.get("signals", [])}
    title = lead.get("title", "")
    company = format_company_name(lead.get("company_name"))
    stage = (lead.get("company_stage") or "").lower()
    seniority = SENIORITY_DISPLAY.get(lead.get("seniority_tier", ""), "VP")

    fragments = []

    if "first_hire" in signals:
        fragments.append(f"First {title} hire — building function from scratch")
    elif "reports_ceo" in signals and "build_team" in signals:
        fragments.append("Reports to CEO with team build mandate")
    elif "reports_ceo" in signals:
        fragments.append("Reports directly to CEO — high-visibility role")
    elif "reports_cro" in signals and "build_team" in signals:
        fragments.append("Reports to CRO with team build mandate")
    elif "build_team" in signals:
        fragments.append("Team build mandate — scaling org")

    if "growth_hire" in signals and not fragments:
        if "series" in stage or "growth" in stage:
            fragments.append(f"Growth hire at {stage.title()} company")
        else:
            fragments.append("Growth hire — expansion role")

    if "immediate" in signals:
        fragments.append("Urgent fill — likely replacing departed leader")

    if "turnaround" in signals:
        fragments.append("Turnaround/transformation mandate")

    if not fragments:
        if "growth_hire" in signals:
            fragments.append("Growth hire")
        else:
            fragments.append(f"{seniority} role with strong signals")

    return " — ".join(fragments[:2])


def get_contextual_stat(lead, geo_data, function_counts=None):
    """Generate a contextual stat line for a lead card.

    Reframes from abundance ('1 of 375') to curation value
    ('Top-scored of 375 screened').
    """
    metro = lead.get("location_metro")
    is_remote = lead.get("is_remote")
    func = lead.get("function_category") or ""
    role_display = FUNCTION_TO_ROLE.get(func, "")

    # For remote roles, use function-specific count if available
    if is_remote or (lead.get("location_type") or "").lower() == "remote":
        if function_counts and func in function_counts:
            return f"Top-scored of {function_counts[func]} remote {role_display or 'VP+'} roles screened"
        for g in geo_data:
            if g["metro"] == "Remote":
                return f"Top-scored of {g['count']} remote VP+ roles screened"

    if metro:
        for g in geo_data:
            if g["metro"] == metro:
                return f"Top-scored of {g['count']} VP+ roles in {metro}"

    return ""


def compute_repost_counts(conn):
    """Compute how many scrape dates each title+company combo appears across.

    A job appearing in multiple scrapes = likely reposted / unfilled.
    Multi-location postings within the SAME scrape are NOT counted as reposts.
    Returns a dict of (normalized_company, title) -> scrape_date_count.
    """
    tier_placeholders = ",".join("?" for _ in VP_TIERS)
    rows = conn.execute(f"""
        SELECT LOWER(company_name_normalized), LOWER(title),
               COUNT(DISTINCT date(date_scraped)) as scrape_count
        FROM jobs
        WHERE seniority_tier IN ({tier_placeholders})
          AND is_active = 1
          AND company_name_normalized IS NOT NULL
        GROUP BY LOWER(company_name_normalized), LOWER(title)
        HAVING scrape_count > 1
    """, VP_TIERS).fetchall()
    return {(r[0], r[1]): r[2] for r in rows}


def deduplicate_leads(leads):
    """Deduplicate leads by title + normalized company name.

    When the same role is posted in multiple locations (common on Indeed),
    keep the highest-scored instance and track how many locations it appears in.
    """
    seen = {}  # (company_normalized, title) -> best lead
    for lead in leads:
        key = (
            (lead.get("company_name_normalized") or "").lower(),
            (lead.get("title") or "").lower(),
        )
        if key not in seen or lead["score"] > seen[key]["score"]:
            seen[key] = lead

    deduped = list(seen.values())
    deduped.sort(key=lambda x: x["score"], reverse=True)
    return deduped


def get_best_job_url(lead):
    """Return the best URL for applying to a job.

    Prefers source_url (direct ATS link - Workday, Lever, Greenhouse, etc.)
    Falls back to company_url if source_url is missing.
    """
    return lead.get("source_url") or lead.get("company_url") or "#"


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 3: EXCEL WORKBOOK GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════


def _xl_cell(ws, row, col, value, font=None, fill=None, alignment=None,
             border=None, number_format=None, hyperlink=None):
    """Set cell value and styles."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    if hyperlink:
        cell.hyperlink = hyperlink
    return cell


def _build_xl_leads(wb, leads, ref_date=None):
    """Sheet 1: Top Leads — scored, color-coded, filterable."""
    ws = wb.active
    ws.title = "Top Leads"
    ws.sheet_properties.tabColor = AMBER

    columns = [
        ("#", 5), ("Score", 8), ("Title", 28), ("Company", 18),
        ("Location", 26), ("Salary Min", 14), ("Salary Max", 14),
        ("Est. Fee", 12), ("Seniority", 12), ("Signals", 44), ("Days Posted", 13),
        ("Flags", 18), ("Apply", 14), ("Company Site", 16), ("Employees", 12),
        ("Industry", 20), ("Signal Note", 54),
    ]

    for i, (_, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for i, (header, _) in enumerate(columns, 1):
        _xl_cell(ws, 1, i, header,
                 font=XL_HEADER_FONT, fill=XL_HEADER_FILL,
                 alignment=XL_CENTER, border=XL_THIN_BORDER)
    ws.row_dimensions[1].height = 32

    now = ref_date or datetime.now()

    for idx, lead in enumerate(leads):
        row = idx + 2
        ws.row_dimensions[row].height = 30

        row_fill = XL_ALT_ROW if idx % 2 == 1 else None

        score = lead["score"]
        if score >= 40:
            score_fill = XL_SCORE_GOLD
            score_font = Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11)
        elif score >= 30:
            score_fill = XL_SCORE_BLUE
            score_font = Font(name="Plus Jakarta Sans", bold=True, color=WHITE, size=11)
        elif score >= 20:
            score_fill = XL_SCORE_LIGHT
            score_font = Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11)
        else:
            score_fill = XL_SCORE_GRAY
            score_font = Font(name="Plus Jakarta Sans", color="555555", size=11)

        # Days since posted
        date_posted = lead.get("date_posted")
        if date_posted:
            try:
                posted_dt = datetime.strptime(str(date_posted)[:10], "%Y-%m-%d")
                days_ago = (now - posted_dt).days
            except ValueError:
                days_ago = 0
        else:
            days_ago = 0

        if days_ago <= 2:
            days_font = Font(name="Plus Jakarta Sans", bold=True, color=RED_FONT, size=10)
        elif days_ago <= 4:
            days_font = Font(name="Plus Jakarta Sans", bold=True, color=AMBER_FONT, size=10)
        else:
            days_font = Font(name="Plus Jakarta Sans", color=GRAY_FONT, size=10)

        seniority = (lead.get("seniority_tier") or "").replace("_", " ").title()
        if "C Level" in seniority or "C-Level" in seniority:
            seniority = "C-Level"
            sen_font = Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=10)
        elif seniority == "Svp":
            seniority = "SVP"
            sen_font = Font(name="Plus Jakarta Sans", bold=True, color="5B8DEF", size=10)
        elif seniority == "Evp":
            seniority = "EVP"
            sen_font = Font(name="Plus Jakarta Sans", bold=True, color="5B8DEF", size=10)
        else:
            if seniority == "Vp":
                seniority = "VP"
            sen_font = XL_BODY_FONT

        # Build signals string (filter reports_cro for non-sales)
        filtered_lead = {**lead, "signals": filter_signals_for_role(lead)}
        signal_parts = []
        hiring_sig = extract_hiring_signal(filtered_lead)
        team_sig = extract_team_structure(filtered_lead)
        if hiring_sig:
            signal_parts.append(hiring_sig)
        if team_sig:
            signal_parts.extend(team_sig.split(", "))
        extras = extract_extra_signals(lead)
        for key in ("segment", "comp", "motion"):
            if key in extras:
                signal_parts.extend(extras[key])
        signals_str = ", ".join(signal_parts)

        note = generate_signal_note(filtered_lead)
        fee = estimate_placement_fee(lead)
        repost_count = lead.get("repost_count", 0)
        repost_parts = []
        if repost_count > 1:
            repost_parts.append(f"REPOSTED {repost_count}x")
        if lead.get("is_search_firm"):
            repost_parts.append("RETAINED SEARCH")
        repost_text = " | ".join(repost_parts)
        repost_font = Font(name="Plus Jakarta Sans", bold=True, color="7C3AED" if lead.get("is_search_firm") else "856404", size=10) if repost_parts else XL_BODY_FONT

        values = [
            (idx + 1, XL_CENTER, XL_BODY_FONT, row_fill),
            (score, XL_CENTER, score_font, score_fill),
            (lead.get("title") or "Untitled", XL_LEFT, XL_BODY_BOLD, row_fill),
            (format_company_name(lead.get("company_name")), XL_LEFT, XL_BODY_FONT, row_fill),
            (clean_location(lead), XL_LEFT, XL_BODY_FONT, row_fill),
            (lead.get("annual_salary_min"), XL_RIGHT, XL_BODY_FONT, row_fill),
            (lead.get("annual_salary_max"), XL_RIGHT, XL_BODY_FONT, row_fill),
            (fee or "", XL_CENTER, Font(name="Plus Jakarta Sans", bold=True, color="2E7D32", size=10), row_fill),
            (seniority, XL_CENTER, sen_font, row_fill),
            (signals_str, XL_LEFT_WRAP, XL_BODY_FONT, row_fill),
            (days_ago, XL_CENTER, days_font, row_fill),
            (repost_text, XL_CENTER, repost_font, row_fill),
            ("Apply", XL_CENTER, XL_LINK_FONT, row_fill),
            ("Website", XL_CENTER, XL_LINK_FONT, row_fill) if lead.get("company_url") else ("", XL_CENTER, XL_BODY_FONT, row_fill),
            (lead.get("company_num_employees") or "", XL_RIGHT, XL_BODY_FONT, row_fill),
            (INDUSTRY_MAP.get(lead.get("company_industry") or "", lead.get("company_industry") or ""), XL_LEFT, XL_BODY_FONT, row_fill),
            (note, XL_LEFT_WRAP, XL_BODY_FONT, row_fill),
        ]

        for col_idx, (val, align, font, fill) in enumerate(values, 1):
            cell = _xl_cell(ws, row, col_idx, val,
                            font=font, fill=fill, alignment=align,
                            border=XL_THIN_BORDER)
            if col_idx in (6, 7) and val:
                cell.number_format = '$#,##0'

        # Apply hyperlink (col 13)
        job_url = get_best_job_url(lead)
        if job_url and job_url != "#":
            apply_cell = ws.cell(row=row, column=13)
            apply_cell.hyperlink = job_url
            apply_cell.value = "Apply"
            apply_cell.font = XL_LINK_FONT

        # Company website hyperlink (col 14)
        company_url = lead.get("company_url") or ""
        if company_url:
            co_cell = ws.cell(row=row, column=14)
            co_cell.hyperlink = company_url
            co_cell.value = "Website"
            co_cell.font = XL_LINK_FONT

    ws.auto_filter.ref = f"A1:Q{len(leads) + 1}"
    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 100


def _build_xl_intel(wb, analytics, leads, date_str):
    """Sheet 2: Market Intel — salary benchmarks, velocity, companies, geo."""
    ws = wb.create_sheet("Market Intel")
    ws.sheet_properties.tabColor = "5B8DEF"

    col_widths = {1: 4, 2: 30, 3: 14, 4: 14, 5: 14, 6: 16, 7: 4, 8: 30, 9: 14, 10: 14}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    r = 1
    total_leads = len(leads)

    # Title bar
    ws.merge_cells(f"B{r}:F{r}")
    title_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    _xl_cell(ws, r, 2,
             f"ExecSignals  \u2014  Market Intelligence Brief  |  Week of {date_str}",
             font=Font(name="DM Serif Display", bold=True, color=AMBER, size=14),
             fill=title_fill,
             alignment=Alignment(horizontal="left", vertical="center"))
    for c in range(3, 7):
        _xl_cell(ws, r, c, None, fill=title_fill)
    ws.row_dimensions[r].height = 42
    r += 1

    # Subtitle
    ws.merge_cells(f"B{r}:F{r}")
    _xl_cell(ws, r, 2,
             f"VP+ hiring intelligence  |  {total_leads} new VP+ roles this week",
             font=Font(name="Plus Jakarta Sans", italic=True, color=GRAY_FONT, size=9),
             alignment=XL_LEFT)
    ws.row_dimensions[r].height = 20
    r += 2

    # ── SALARY BENCHMARKS ──
    benchmarks = analytics["salary_benchmarks"]
    ws.merge_cells(f"B{r}:F{r}")
    _xl_cell(ws, r, 2, "  SALARY BENCHMARKS \u2014 VP+ ROLES",
             font=Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=12),
             fill=XL_SECTION_FILL,
             alignment=Alignment(horizontal="left", vertical="center"))
    for c in range(3, 7):
        _xl_cell(ws, r, c, None, fill=XL_SECTION_FILL)
    ws.row_dimensions[r].height = 32
    r += 1

    for i, h in enumerate(["Role", "P25", "Median", "P75", "4-Week Trend"]):
        _xl_cell(ws, r, i + 2, h, font=XL_HEADER_FONT, fill=XL_HEADER_FILL,
                 alignment=XL_CENTER, border=XL_THIN_BORDER)
    ws.row_dimensions[r].height = 28
    r += 1

    for idx, b in enumerate(benchmarks):
        fill = XL_ALT_ROW if idx % 2 == 1 else None
        trend = b["trend_pct"]
        if trend > 0:
            trend_font = Font(name="Plus Jakarta Sans", bold=True, color=GREEN_FONT, size=10)
            trend_display = f"\u25B2 +{trend}%"
        elif trend < 0:
            trend_font = Font(name="Plus Jakarta Sans", bold=True, color=RED_FONT, size=10)
            trend_display = f"\u25BC {trend}%"
        else:
            trend_font = XL_BODY_FONT
            trend_display = f"\u25AC 0%"

        for ci, (v, al, fnt) in enumerate([
            (b["role"], XL_LEFT, XL_BODY_BOLD),
            (b["p25"], XL_RIGHT, XL_BODY_FONT),
            (b["median"], XL_RIGHT, Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11)),
            (b["p75"], XL_RIGHT, XL_BODY_FONT),
            (trend_display, XL_CENTER, trend_font),
        ]):
            _xl_cell(ws, r, ci + 2, v, font=fnt, fill=fill, alignment=al, border=XL_THIN_BORDER)
        ws.row_dimensions[r].height = 26
        r += 1

    r += 2

    # ── HIRING VELOCITY ──
    velocity = analytics["industry_velocity"]
    ws.merge_cells(f"B{r}:F{r}")
    _xl_cell(ws, r, 2, "  HIRING VELOCITY BY INDUSTRY",
             font=Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=12),
             fill=XL_SECTION_FILL,
             alignment=Alignment(horizontal="left", vertical="center"))
    for c in range(3, 7):
        _xl_cell(ws, r, c, None, fill=XL_SECTION_FILL)
    ws.row_dimensions[r].height = 32
    r += 1

    for i, h in enumerate(["Industry", "VP+ Openings This Week", "WoW Change", "", ""]):
        _xl_cell(ws, r, i + 2, h, font=XL_HEADER_FONT, fill=XL_HEADER_FILL,
                 alignment=XL_CENTER, border=XL_THIN_BORDER)
    ws.row_dimensions[r].height = 28
    r += 1

    for idx, v in enumerate(velocity):
        fill = XL_ALT_ROW if idx % 2 == 1 else None
        wow = v["wow_pct"]
        if wow > 0:
            wow_font = Font(name="Plus Jakarta Sans", bold=True, color=GREEN_FONT, size=10)
            wow_display = f"\u25B2 +{wow}%"
        elif wow < 0:
            wow_font = Font(name="Plus Jakarta Sans", bold=True, color=RED_FONT, size=10)
            wow_display = f"\u25BC {wow}%"
        else:
            wow_font = XL_BODY_FONT
            wow_display = "0%"

        _xl_cell(ws, r, 2, v["industry"], font=XL_BODY_BOLD, fill=fill,
                 alignment=XL_LEFT, border=XL_THIN_BORDER)
        _xl_cell(ws, r, 3, v["count"],
                 font=Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11),
                 fill=fill, alignment=XL_CENTER, border=XL_THIN_BORDER)
        _xl_cell(ws, r, 4, wow_display, font=wow_font, fill=fill,
                 alignment=XL_CENTER, border=XL_THIN_BORDER)
        _xl_cell(ws, r, 5, None, fill=fill, border=XL_THIN_BORDER)
        _xl_cell(ws, r, 6, None, fill=fill, border=XL_THIN_BORDER)
        ws.row_dimensions[r].height = 26
        r += 1

    r += 2

    # ── TOP COMPANIES + GEO (side by side) ──
    companies = analytics["top_companies"]
    geo = analytics["geo_breakdown"]

    ws.merge_cells(f"B{r}:D{r}")
    _xl_cell(ws, r, 2, "  TOP HIRING COMPANIES",
             font=Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=12),
             fill=XL_SECTION_FILL,
             alignment=Alignment(horizontal="left", vertical="center"))
    for c in [3, 4]:
        _xl_cell(ws, r, c, None, fill=XL_SECTION_FILL)

    ws.merge_cells(f"F{r}:H{r}")
    _xl_cell(ws, r, 6, "  GEO BREAKDOWN",
             font=Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=12),
             fill=XL_SECTION_FILL,
             alignment=Alignment(horizontal="left", vertical="center"))
    for c in [7, 8]:
        _xl_cell(ws, r, c, None, fill=XL_SECTION_FILL)
    ws.column_dimensions["H"].width = 14
    ws.row_dimensions[r].height = 32
    r += 1

    for i, h in enumerate(["Company", "VP+ Roles", ""]):
        _xl_cell(ws, r, i + 2, h, font=XL_HEADER_FONT, fill=XL_HEADER_FILL,
                 alignment=XL_CENTER, border=XL_THIN_BORDER)
    for i, h in enumerate(["Metro Area", "VP+ Roles", "WoW"]):
        _xl_cell(ws, r, i + 6, h, font=XL_HEADER_FONT, fill=XL_HEADER_FILL,
                 alignment=XL_CENTER, border=XL_THIN_BORDER)
    ws.row_dimensions[r].height = 28
    r += 1

    max_rows = max(len(companies), len(geo))
    for idx in range(max_rows):
        fill = XL_ALT_ROW if idx % 2 == 1 else None
        ws.row_dimensions[r].height = 26

        if idx < len(companies):
            comp = companies[idx]
            name = format_company_name(comp["company"])
            if comp["is_new"]:
                name += " \u2605"  # star for new
            _xl_cell(ws, r, 2, name, font=XL_BODY_BOLD, fill=fill,
                     alignment=XL_LEFT, border=XL_THIN_BORDER)
            _xl_cell(ws, r, 3, comp["count"],
                     font=Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11),
                     fill=fill, alignment=XL_CENTER, border=XL_THIN_BORDER)
            _xl_cell(ws, r, 4, None, fill=fill, border=XL_THIN_BORDER)

        if idx < len(geo):
            g = geo[idx]
            _xl_cell(ws, r, 6, g["metro"], font=XL_BODY_BOLD, fill=fill,
                     alignment=XL_LEFT, border=XL_THIN_BORDER)
            _xl_cell(ws, r, 7, g["count"],
                     font=Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11),
                     fill=fill, alignment=XL_CENTER, border=XL_THIN_BORDER)
            wow = g["wow_pct"]
            if wow > 0:
                wow_font = Font(name="Plus Jakarta Sans", color=GREEN_FONT, size=10)
                wow_d = f"+{wow}%"
            elif wow < 0:
                wow_font = Font(name="Plus Jakarta Sans", color=RED_FONT, size=10)
                wow_d = f"{wow}%"
            else:
                wow_font = Font(name="Plus Jakarta Sans", color=GRAY_FONT, size=10)
                wow_d = "0%"
            _xl_cell(ws, r, 8, wow_d, font=wow_font, fill=fill,
                     alignment=XL_CENTER, border=XL_THIN_BORDER)

        r += 1

    r += 2

    # ── KEY TAKEAWAYS (auto-generated) ──
    ws.merge_cells(f"B{r}:H{r}")
    _xl_cell(ws, r, 2, "  THIS WEEK'S KEY TAKEAWAYS",
             font=Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=12),
             fill=XL_SECTION_FILL,
             alignment=Alignment(horizontal="left", vertical="center"))
    for c in range(3, 9):
        _xl_cell(ws, r, c, None, fill=XL_SECTION_FILL)
    ws.row_dimensions[r].height = 32
    r += 1

    takeaways = _generate_takeaways(leads, analytics)
    for tk in takeaways:
        ws.merge_cells(f"B{r}:H{r}")
        _xl_cell(ws, r, 2, f"\u25CF  {tk}",
                 font=Font(name="Plus Jakarta Sans", color=DARK_TEXT, size=10),
                 alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
        for c in range(3, 9):
            _xl_cell(ws, r, c, None)
        ws.row_dimensions[r].height = 24
        r += 1

    r += 2

    # Footer
    ws.merge_cells(f"B{r}:H{r}")
    _xl_cell(ws, r, 2,
             f"ExecSignals  |  The Monday Brief  |  execsignals.com  |  {total_leads} VP+ roles scored",
             font=Font(name="Plus Jakarta Sans", italic=True, color=GRAY_FONT, size=9),
             alignment=XL_LEFT)
    r += 1
    ws.merge_cells(f"B{r}:H{r}")
    _xl_cell(ws, r, 2,
             "Confidential \u2014 for subscriber use only. Do not redistribute.",
             font=Font(name="Plus Jakarta Sans", italic=True, color=RED_FONT, size=9),
             alignment=XL_LEFT)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _generate_takeaways(leads, analytics):
    """Auto-generate key takeaways from analytics data."""
    takeaways = []

    # Top velocity industry
    velocity = analytics.get("industry_velocity", [])
    if velocity:
        top = velocity[0]
        takeaways.append(
            f"{top['industry']} leads VP+ hiring with {top['count']} openings "
            f"({top['wow_display']} WoW)"
        )

    # Salary highlight
    benchmarks = analytics.get("salary_benchmarks", [])
    if benchmarks:
        highest = max(benchmarks, key=lambda x: x.get("median_raw", 0))
        takeaways.append(
            f"{highest['role']} has the highest median salary at {highest['median']} "
            f"({highest['trend_display']} 4-week trend)"
        )

    # C-level count
    c_level_count = sum(1 for l in leads if l.get("seniority_tier") == "c_level")
    if c_level_count >= 2:
        takeaways.append(
            f"{c_level_count} C-Level roles posted this week — "
            f"retained search opportunities"
        )

    # Top geo
    geo = analytics.get("geo_breakdown", [])
    if geo:
        top_geo = geo[0]
        takeaways.append(
            f"{top_geo['metro']} leads with {top_geo['count']} VP+ openings "
            f"({top_geo['wow_display']} WoW)"
        )

    # Build Team signal prevalence
    build_team_count = sum(
        1 for l in leads
        if any(s["signal_id"] == "build_team" for s in l.get("signals", []))
    )
    if build_team_count >= 3:
        pct = round(build_team_count / len(leads) * 100)
        takeaways.append(
            f"\"Build Team\" signal in {build_team_count} of {len(leads)} top leads "
            f"({pct}%) — companies investing in org growth"
        )

    # Stage distribution insight
    stages = analytics.get("company_stage", [])
    enterprise_pct = next((s["pct"] for s in stages if "Enterprise" in s["stage"]), 0)
    if enterprise_pct > 0:
        takeaways.append(
            f"Enterprise/Public companies account for {enterprise_pct}% of VP+ hiring"
        )

    return takeaways[:6]


def generate_excel(leads, analytics, output_path, date_str, ref_date=None):
    """Generate the full Excel workbook."""
    wb = openpyxl.Workbook()
    _build_xl_leads(wb, leads, ref_date=ref_date)
    _build_xl_intel(wb, analytics, leads, date_str)
    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 4: PDF ONE-PAGER GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════


def generate_market_intel_html(analytics, summary, date_str):
    """Generate print-ready Market Intel one-pager HTML."""

    # Salary benchmarks rows
    salary_rows = ""
    for b in analytics["salary_benchmarks"]:
        trend = b["trend_pct"]
        if trend > 0:
            trend_class = "trend-up"
            trend_text = f"&#9650; {trend}%"
        elif trend < 0:
            trend_class = "trend-down"
            trend_text = f"&#9660; {trend}%"
        else:
            trend_class = "trend-flat"
            trend_text = "&#9644; 0%"
        salary_rows += f"""
            <tr>
              <td class="role-cell">{html.escape(b['role'])}</td>
              <td>{b['p25']}</td>
              <td>{b['median']}</td>
              <td>{b['p75']}</td>
              <td class="{trend_class}">{trend_text}</td>
            </tr>"""

    # Industry velocity rows
    velocity_rows = ""
    for v in analytics["industry_velocity"][:6]:
        wow = v["wow_pct"]
        cls = "change-positive" if wow > 0 else ("change-negative" if wow < 0 else "")
        sign = "+" if wow > 0 else ""
        velocity_rows += f"""
            <tr>
              <td class="role-cell">{html.escape(v['industry'])}</td>
              <td>{v['count']}</td>
              <td class="{cls}">{sign}{wow}%</td>
            </tr>"""

    # Top companies rows
    company_rows = ""
    for comp in analytics["top_companies"][:10]:
        badge = '<span class="badge-new">new</span>' if comp["is_new"] else ""
        company_rows += f"""
            <tr>
              <td class="role-cell">{html.escape(format_company_name(comp['company']))}{badge}</td>
              <td>{comp['count']}</td>
            </tr>"""

    # Geo rows
    geo_rows = ""
    for g in analytics["geo_breakdown"][:8]:
        wow = g["wow_pct"]
        cls = "change-positive" if wow > 0 else ("change-negative" if wow < 0 else "")
        sign = "+" if wow > 0 else ""
        geo_rows += f"""
            <tr>
              <td class="role-cell">{html.escape(g['metro'])}</td>
              <td>{g['count']}</td>
              <td class="{cls}">{sign}{wow}%</td>
            </tr>"""

    # Company stage bar segments
    stage_segments = ""
    for s in analytics["company_stage"]:
        if s["pct"] < 3:
            continue
        css_class = s["stage"].lower().split("/")[0].strip().split(" ")[0]
        stage_segments += f'<div class="stage-segment {css_class}" style="width: {s["pct"]}%;">{s["pct"]}%</div>\n'

    # Unique companies count
    unique_companies = len(set(
        (analytics.get("top_companies") or [{}])[0].get("company", "")
        for _ in range(1)
    ))

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ExecSignals Market Intelligence Brief — {html.escape(date_str)}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=IBM+Plex+Mono:wght@400;500;600&family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
  *, *::before, *::after {{ margin: 0; padding: 0; box-sizing: border-box; }}
  :root {{
    --amber: #D4A054; --amber-light: #F5E6CC; --amber-bg: rgba(212, 160, 84, 0.08);
    --navy: #0C0F1A; --navy-soft: #1A1E2E; --white: #FFFFFF;
    --gray-50: #F9FAFB; --gray-100: #F3F4F6; --gray-200: #E5E7EB;
    --gray-300: #D1D5DB; --gray-400: #9CA3AF; --gray-500: #6B7280;
    --gray-600: #4B5563; --gray-700: #374151; --gray-800: #1F2937;
    --green: #16A34A; --green-light: #DCFCE7; --red: #DC2626; --red-light: #FEE2E2;
    --font-display: 'DM Serif Display', serif;
    --font-body: 'Plus Jakarta Sans', sans-serif;
    --font-mono: 'IBM Plex Mono', monospace;
  }}
  html, body {{ font-family: var(--font-body); background: #E5E7EB; color: var(--gray-800); -webkit-font-smoothing: antialiased; }}
  .page {{ width: 8.5in; height: 11in; margin: 0.5in auto; background: var(--white); box-shadow: 0 4px 24px rgba(0,0,0,0.12); display: grid; grid-template-rows: auto auto 1fr auto auto; overflow: hidden; position: relative; }}
  .header {{ background: var(--navy); padding: 14px 28px; display: flex; align-items: center; justify-content: space-between; gap: 16px; }}
  .header-left {{ display: flex; align-items: baseline; gap: 16px; }}
  .logo {{ font-family: var(--font-display); font-size: 20px; color: var(--amber); letter-spacing: 0.02em; }}
  .header-title {{ font-family: var(--font-body); font-size: 12px; font-weight: 500; color: rgba(255,255,255,0.85); letter-spacing: 0.06em; text-transform: uppercase; }}
  .header-date {{ font-family: var(--font-mono); font-size: 11px; color: var(--gray-400); white-space: nowrap; }}
  .stats-row {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 1px; background: var(--gray-200); border-bottom: 1px solid var(--gray-200); }}
  .stat-box {{ background: var(--white); padding: 14px 20px; text-align: center; }}
  .stat-value {{ font-family: var(--font-mono); font-size: 26px; font-weight: 600; color: var(--navy); line-height: 1.1; }}
  .stat-label {{ font-family: var(--font-body); font-size: 10px; font-weight: 600; color: var(--gray-500); text-transform: uppercase; letter-spacing: 0.08em; margin-top: 4px; }}
  .content {{ padding: 16px 28px 12px; display: grid; grid-template-rows: auto auto auto; gap: 14px; }}
  .two-col {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
  .section-title {{ font-family: var(--font-display); font-size: 13px; color: var(--navy); margin-bottom: 8px; padding-bottom: 5px; border-bottom: 2px solid var(--amber); display: flex; align-items: center; gap: 6px; }}
  .section-title::before {{ content: ''; display: inline-block; width: 3px; height: 13px; background: var(--amber); border-radius: 1px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 10px; }}
  table th {{ font-family: var(--font-mono); font-size: 8.5px; font-weight: 600; color: var(--gray-400); text-transform: uppercase; letter-spacing: 0.1em; text-align: right; padding: 3px 6px 4px; border-bottom: 1px solid var(--gray-200); }}
  table th:first-child {{ text-align: left; }}
  table td {{ padding: 4px 6px; border-bottom: 1px solid var(--gray-100); font-family: var(--font-body); font-size: 10px; color: var(--gray-700); }}
  table td:not(:first-child) {{ text-align: right; font-family: var(--font-mono); font-size: 10px; }}
  table tr:last-child td {{ border-bottom: none; }}
  .role-cell {{ font-weight: 600; color: var(--navy); }}
  .trend-up {{ color: var(--green); font-size: 9px; }}
  .trend-down {{ color: var(--red); font-size: 9px; }}
  .trend-flat {{ color: var(--gray-400); font-size: 9px; }}
  .badge-new {{ font-family: var(--font-mono); font-size: 7.5px; font-weight: 600; color: var(--green); background: var(--green-light); padding: 1px 5px; border-radius: 3px; letter-spacing: 0.04em; text-transform: uppercase; margin-left: 4px; vertical-align: middle; }}
  .change-positive {{ color: var(--green); }}
  .change-negative {{ color: var(--red); }}
  .stage-section {{ padding: 0 28px 10px; }}
  .stage-title {{ font-family: var(--font-display); font-size: 13px; color: var(--navy); margin-bottom: 8px; padding-bottom: 5px; border-bottom: 2px solid var(--amber); display: flex; align-items: center; gap: 6px; }}
  .stage-title::before {{ content: ''; display: inline-block; width: 3px; height: 13px; background: var(--amber); border-radius: 1px; }}
  .stage-bar-container {{ margin-bottom: 6px; }}
  .stage-bar {{ display: flex; height: 26px; border-radius: 4px; overflow: hidden; gap: 1px; }}
  .stage-segment {{ display: flex; align-items: center; justify-content: center; font-family: var(--font-mono); font-size: 9px; font-weight: 600; color: var(--white); }}
  .stage-segment.enterprise {{ background: var(--navy); }}
  .stage-segment.late {{ background: #374151; }}
  .stage-segment.growth {{ background: var(--amber); color: var(--navy); }}
  .stage-segment.early {{ background: #D4A054aa; color: var(--navy); }}
  .stage-segment.unknown {{ background: var(--gray-300); color: var(--gray-600); }}
  .stage-legend {{ display: flex; gap: 16px; justify-content: center; }}
  .stage-legend-item {{ display: flex; align-items: center; gap: 5px; font-family: var(--font-body); font-size: 9px; color: var(--gray-500); }}
  .stage-legend-dot {{ width: 8px; height: 8px; border-radius: 2px; }}
  .stage-legend-dot.enterprise {{ background: var(--navy); }}
  .stage-legend-dot.late {{ background: #374151; }}
  .stage-legend-dot.growth {{ background: var(--amber); }}
  .stage-legend-dot.early {{ background: #D4A054aa; }}
  .stage-legend-dot.unknown {{ background: var(--gray-300); }}
  .footer {{ background: var(--gray-50); border-top: 1px solid var(--gray-200); padding: 10px 28px; display: flex; justify-content: space-between; align-items: center; }}
  .footer-left {{ font-family: var(--font-body); font-size: 9px; color: var(--gray-500); }}
  .footer-left strong {{ color: var(--amber); font-family: var(--font-display); font-weight: normal; font-size: 10px; }}
  .footer-right {{ font-family: var(--font-mono); font-size: 8px; color: var(--gray-400); text-transform: uppercase; letter-spacing: 0.06em; }}
  @media print {{
    html, body {{ background: white; margin: 0; padding: 0; }}
    .page {{ width: 100%; height: 100%; margin: 0; box-shadow: none; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    @page {{ size: letter; margin: 0; }}
  }}
</style>
</head>
<body>
<div class="page">
  <div class="header">
    <div class="header-left">
      <span class="logo">ExecSignals</span>
      <span class="header-title">Market Intelligence Brief</span>
    </div>
    <span class="header-date">{html.escape(date_str)}</span>
  </div>
  <div class="stats-row">
    <div class="stat-box">
      <div class="stat-value">{summary['total']}</div>
      <div class="stat-label">VP+ Leads Scored</div>
    </div>
    <div class="stat-box">
      <div class="stat-value">{summary['avg_salary']}</div>
      <div class="stat-label">Avg Max Salary</div>
    </div>
    <div class="stat-box">
      <div class="stat-value">{len(analytics['top_companies'])}</div>
      <div class="stat-label">Companies Hiring</div>
    </div>
    <div class="stat-box">
      <div class="stat-value">{summary['growth_pct']}%</div>
      <div class="stat-label">Growth Hires</div>
    </div>
  </div>
  <div class="content">
    <div class="two-col">
      <div>
        <div class="section-title">Salary Benchmarks</div>
        <table>
          <thead><tr><th>Role</th><th>P25</th><th>Median</th><th>P75</th><th>4wk</th></tr></thead>
          <tbody>{salary_rows}</tbody>
        </table>
      </div>
      <div>
        <div class="section-title">Hiring Velocity by Industry</div>
        <table>
          <thead><tr><th>Industry</th><th>Leads</th><th>WoW</th></tr></thead>
          <tbody>{velocity_rows}</tbody>
        </table>
      </div>
    </div>
    <div class="two-col">
      <div>
        <div class="section-title">Top Hiring Companies</div>
        <table>
          <thead><tr><th>Company</th><th>Open VP+ Roles</th></tr></thead>
          <tbody>{company_rows}</tbody>
        </table>
      </div>
      <div>
        <div class="section-title">VP+ Leads by Metro</div>
        <table>
          <thead><tr><th>Metro</th><th>Leads</th><th>WoW</th></tr></thead>
          <tbody>{geo_rows}</tbody>
        </table>
      </div>
    </div>
  </div>
  <div class="stage-section">
    <div class="stage-title">Company Stage Distribution</div>
    <div class="stage-bar-container">
      <div class="stage-bar">
        {stage_segments}
      </div>
    </div>
    <div class="stage-legend">
      <div class="stage-legend-item"><div class="stage-legend-dot enterprise"></div>Enterprise</div>
      <div class="stage-legend-item"><div class="stage-legend-dot late"></div>Late Stage</div>
      <div class="stage-legend-item"><div class="stage-legend-dot growth"></div>Growth</div>
      <div class="stage-legend-item"><div class="stage-legend-dot early"></div>Early Stage</div>
      <div class="stage-legend-item"><div class="stage-legend-dot unknown"></div>Unknown</div>
    </div>
  </div>
  <div class="footer">
    <div class="footer-left">
      <strong>ExecSignals</strong> &mdash; The Monday Brief&ensp;|&ensp;execsignals.com&ensp;|&ensp;Pariter Media Inc.
    </div>
    <div class="footer-right">For client use. Updated weekly.</div>
  </div>
</div>
</body>
</html>"""


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 5: EMAIL GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════


def generate_email_html(leads, analytics, summary, date_str, date_range, ref_date=None, file_date=None):
    """Generate The Monday Brief email HTML."""
    now = ref_date or datetime.now()
    fd = file_date or datetime.now().strftime("%b%d")
    xlsx_name = f"ExecSignals_{fd}.xlsx"
    pdf_name = f"MarketIntel_{fd}.html"

    # ── Summary stats bar ──
    seniority_html = ""
    for tier, count in summary["seniority"].items():
        seniority_html += f'{tier}: <strong style="color:#0C0F1A;">{count}</strong> &middot;\n'
    seniority_html = seniority_html.rstrip(" &middot;\n")

    segment_html = ""
    for seg, pct in summary["segment"].items():
        segment_html += f'{seg}: <strong style="color:#0C0F1A;">{pct}%</strong> &middot;\n'
    segment_html = segment_html.rstrip(" &middot;\n")

    # ── Top 5 full lead cards ──
    top5_html = ""
    for i, lead in enumerate(leads[:5], 1):
        title = html.escape(lead.get("title") or "Untitled")
        company = html.escape(format_company_name(lead.get("company_name")))
        salary = html.escape(format_salary(lead.get("annual_salary_min"), lead.get("annual_salary_max")))
        location = html.escape(clean_location(lead))
        job_url = html.escape(get_best_job_url(lead))
        score = lead["score"]
        seniority = SENIORITY_DISPLAY.get(lead.get("seniority_tier", ""), "VP")

        # Company website link
        company_url = lead.get("company_url") or ""
        company_link_html = ""
        if company_url:
            safe_co_url = html.escape(company_url)
            company_link_html = f' &middot; <a href="{safe_co_url}" style="color:#888;font-size:11px;text-decoration:none;">Company &rsaquo;</a>'

        # Days ago
        date_posted = lead.get("date_posted")
        if date_posted:
            try:
                posted_dt = datetime.strptime(str(date_posted)[:10], "%Y-%m-%d")
                days_ago = (now - posted_dt).days
            except ValueError:
                days_ago = 0
        else:
            days_ago = 0

        if days_ago <= 2:
            days_color = "#EF4444"
        elif days_ago <= 4:
            days_color = "#D4A054"
        else:
            days_color = "#888"
        days_text = f"POSTED {days_ago} DAY{'S' if days_ago != 1 else ''} AGO"

        employees = lead.get("company_num_employees") or ""
        emp_text = f" &middot; ~{employees} emp" if employees else ""
        industry_raw = lead.get("company_industry") or ""
        industry = INDUSTRY_MAP.get(industry_raw, industry_raw)
        ind_text = f" &middot; {html.escape(industry)}" if industry else ""

        # Repost badge
        repost_count = lead.get("repost_count", 0)
        repost_badge = ""
        if repost_count > 1:
            repost_badge = f'<span style="display:inline-block;background:#FFF3CD;color:#856404;padding:2px 7px;border-radius:3px;font-size:10px;font-weight:700;margin-right:3px;">REPOSTED {repost_count}x</span>'

        # Retained search badge
        if lead.get("is_search_firm"):
            repost_badge += '<span style="display:inline-block;background:#F3E8FF;color:#7C3AED;padding:2px 7px;border-radius:3px;font-size:10px;font-weight:700;margin-right:3px;">RETAINED SEARCH</span>'

        # Signal badges (filter reports_cro for non-sales roles)
        filtered_lead = {**lead, "signals": filter_signals_for_role(lead)}
        signal_badges = ""
        hiring_sig = extract_hiring_signal(filtered_lead)
        team_sig = extract_team_structure(filtered_lead)
        if hiring_sig:
            signal_badges += f'<span style="display:inline-block;background:#E8F5E9;color:#2E7D32;padding:2px 7px;border-radius:3px;font-size:10px;font-weight:700;margin-right:3px;">{html.escape(hiring_sig.upper())}</span>'
        if team_sig:
            for ts in team_sig.split(", "):
                signal_badges += f'<span style="display:inline-block;background:#E3F2FD;color:#1565C0;padding:2px 7px;border-radius:3px;font-size:10px;font-weight:700;margin-right:3px;">{html.escape(ts.upper())}</span>'

        # Placement fee estimate
        fee = estimate_placement_fee(lead)
        fee_html = f' &middot; <span style="color:#06D6A0;font-weight:600;">~{html.escape(fee)} fee</span>' if fee else ""

        note = html.escape(generate_signal_note(filtered_lead))
        ctx_stat = html.escape(get_contextual_stat(lead, analytics["geo_breakdown"], analytics.get("remote_function_counts")))

        top5_html += f"""
                    <tr>
                        <td style="padding:14px 32px;border-bottom:1px solid #f0f0f0;">
                            <div style="font-size:11px;color:#D4A054;font-weight:700;margin-bottom:2px;">#{i} &middot; <span style="color:{days_color};">{days_text}</span></div>
                            <a href="{job_url}" style="color:#0C0F1A;font-size:15px;font-weight:600;text-decoration:underline;text-decoration-color:#D4A054;">{title}</a>
                            <div style="color:#555;font-size:13px;margin-top:2px;">{company} &middot; {location} <span style="color:#aaa;font-size:11px;">{emp_text}{ind_text}</span>{company_link_html}</div>
                            <div style="color:#888;font-size:12px;margin-top:2px;">{salary} &middot; {seniority}{fee_html}</div>
                            <div style="margin-top:6px;">{repost_badge}{signal_badges}</div>
                            <div style="font-size:12px;color:#666;font-style:italic;margin-top:6px;padding-left:10px;border-left:2px solid #F0DFC0;">{note}</div>
                            <div style="font-size:11px;color:#999;margin-top:4px;">{ctx_stat}</div>
                        </td>
                    </tr>"""

    # ── Leads 6-10 compact ──
    compact_html = ""
    for i, lead in enumerate(leads[5:10], 6):
        title = html.escape(lead.get("title") or "Untitled")
        company = html.escape(format_company_name(lead.get("company_name")))
        salary = html.escape(format_salary(lead.get("annual_salary_min"), lead.get("annual_salary_max")))
        location = html.escape(clean_location(lead))
        job_url = html.escape(get_best_job_url(lead))
        score = lead["score"]
        fee = estimate_placement_fee(lead)
        fee_text = f" &middot; ~{html.escape(fee)}" if fee else ""

        date_posted = lead.get("date_posted")
        if date_posted:
            try:
                posted_dt = datetime.strptime(str(date_posted)[:10], "%Y-%m-%d")
                days_ago = (now - posted_dt).days
            except ValueError:
                days_ago = 0
        else:
            days_ago = 0

        if days_ago <= 2:
            days_color = "#EF4444"
        elif days_ago <= 4:
            days_color = "#D4A054"
        else:
            days_color = "#888"

        metro = lead.get("location_metro") or lead.get("location_state") or ""
        if lead.get("is_remote"):
            metro = "Remote"
        metro_html = f" &middot; {html.escape(metro)}" if metro else ""

        repost_count = lead.get("repost_count", 0)
        repost_tag = f' <span style="background:#FFF3CD;color:#856404;padding:1px 5px;border-radius:2px;font-size:9px;font-weight:700;">{repost_count}x</span>' if repost_count > 1 else ""
        if lead.get("is_search_firm"):
            repost_tag += ' <span style="background:#F3E8FF;color:#7C3AED;padding:1px 5px;border-radius:2px;font-size:9px;font-weight:700;">RETAINED</span>'

        # Compact signal hints
        filtered_c = {**lead, "signals": filter_signals_for_role(lead)}
        sig_parts = []
        h_sig = extract_hiring_signal(filtered_c)
        t_sig = extract_team_structure(filtered_c)
        if h_sig:
            sig_parts.append(h_sig)
        if t_sig:
            sig_parts.extend(t_sig.split(", ")[:2])
        signal_hint = ""
        if sig_parts:
            hint_text = html.escape(" | ".join(sig_parts[:2]))
            signal_hint = f' <span style="color:#999;font-size:9px;font-weight:500;">[{hint_text}]</span>'

        compact_html += f"""
                                <tr style="border-bottom:1px solid #f0f0f0;">
                                    <td style="padding:8px 0;color:#D4A054;font-weight:700;width:60px;">#{i}</td>
                                    <td style="padding:8px 0;"><a href="{job_url}" style="color:#0C0F1A;text-decoration:underline;text-decoration-color:#D4A054;font-weight:600;">{title}</a>{repost_tag}{signal_hint} &middot; {company}{metro_html} &middot; {salary}{fee_text} <span style="color:{days_color};font-size:10px;font-weight:600;">&middot; {days_ago}d</span></td>
                                </tr>"""

    # ── Salary benchmarks table ──
    salary_table = ""
    for b in analytics["salary_benchmarks"]:
        trend = b["trend_pct"]
        if trend > 0:
            trend_color = "#06D6A0"
            trend_arrow = "&#9650;"
            trend_text = f"+{trend}%"
        elif trend < 0:
            trend_color = "#EF4444"
            trend_arrow = "&#9660;"
            trend_text = f"{trend}%"
        else:
            trend_color = "#888"
            trend_arrow = "&#9644;"
            trend_text = "0%"

        salary_table += f"""
                                <tr style="border-top:1px solid #f0f0f0;">
                                    <td style="padding:7px 0;color:#333;font-weight:500;">{html.escape(b['role'])}</td>
                                    <td align="center" style="padding:7px 0;color:#666;">{b['p25']}</td>
                                    <td align="center" style="padding:7px 0;color:#0C0F1A;font-weight:700;">{b['median']}</td>
                                    <td align="center" style="padding:7px 0;color:#666;">{b['p75']}</td>
                                    <td align="right" style="padding:7px 0;color:{trend_color};font-weight:600;">{trend_arrow} {trend_text}</td>
                                </tr>"""

    # ── Velocity rows ──
    velocity_html = ""
    for v in analytics["industry_velocity"][:6]:
        wow = v["wow_pct"]
        if wow > 0:
            wow_color = "#06D6A0"
            wow_text = f"+{wow}%"
        elif wow < 0:
            wow_color = "#EF4444"
            wow_text = f"{wow}%"
        else:
            wow_color = "#888"
            wow_text = "0%"

        velocity_html += f"""
                                            <tr style="border-bottom:1px solid #f5f5f5;">
                                                <td style="padding:5px 0;color:#333;">{html.escape(v['industry'])}</td>
                                                <td align="right" style="padding:5px 0;color:{wow_color};font-weight:700;">{wow_text} <span style="color:#888;font-weight:400;">({v['count']})</span></td>
                                            </tr>"""

    # ── Companies rows ──
    companies_html = ""
    new_section = False
    for comp in analytics["top_companies"]:
        if comp["is_new"] and not new_section:
            companies_html += f"""
                                            <tr style="border-bottom:1px solid #f5f5f5;">
                                                <td style="padding:5px 0;color:#333;"><em>New this week:</em></td>
                                                <td style="padding:5px 0;"></td>
                                            </tr>"""
            new_section = True

        comp_name = format_company_name(comp['company'])
        if comp["is_new"]:
            companies_html += f"""
                                            <tr style="border-bottom:1px solid #f5f5f5;">
                                                <td style="padding:5px 0;color:#06D6A0;font-weight:500;">{html.escape(comp_name)}</td>
                                                <td align="right" style="padding:5px 0;font-weight:600;color:#06D6A0;">{comp['count']} roles</td>
                                            </tr>"""
        else:
            companies_html += f"""
                                            <tr style="border-bottom:1px solid #f5f5f5;">
                                                <td style="padding:5px 0;color:#333;">{html.escape(comp_name)}</td>
                                                <td align="right" style="padding:5px 0;font-weight:600;color:#0C0F1A;">{comp['count']} roles</td>
                                            </tr>"""

    # ── Geo rows ──
    geo_html = ""
    for g in analytics["geo_breakdown"][:8]:
        wow = g["wow_pct"]
        if wow > 0:
            wow_color = "#06D6A0"
            wow_text = f"+{wow}%"
        elif wow < 0:
            wow_color = "#EF4444"
            wow_text = f"{wow}%"
        else:
            wow_color = "#888"
            wow_text = "0%"

        geo_html += f"""
                                            <tr style="border-bottom:1px solid #f5f5f5;">
                                                <td style="padding:5px 0;color:#333;">{html.escape(g['metro'])}</td>
                                                <td align="center" style="padding:5px 0;font-weight:600;color:#0C0F1A;">{g['count']}</td>
                                                <td align="right" style="padding:5px 0;color:{wow_color};font-size:11px;font-weight:600;">{wow_text}</td>
                                            </tr>"""

    # ── Key takeaways (auto-generated from data) ──
    takeaways = []
    # Hottest industry
    if analytics["industry_velocity"]:
        hot = max(analytics["industry_velocity"], key=lambda x: x["wow_pct"])
        if hot["wow_pct"] > 0:
            takeaways.append(f"{hot['industry']} hiring surged <strong>+{hot['wow_pct']}%</strong> WoW ({hot['count']} openings)")
    # Biggest drop
    if analytics["industry_velocity"]:
        cold = min(analytics["industry_velocity"], key=lambda x: x["wow_pct"])
        if cold["wow_pct"] < -10:
            takeaways.append(f"{cold['industry']} down <strong>{cold['wow_pct']}%</strong> WoW")
    # Top salary role
    if analytics["salary_benchmarks"]:
        top_sal = max(analytics["salary_benchmarks"], key=lambda x: x.get("median_raw", 0))
        takeaways.append(f"Highest median: <strong>{top_sal['role']}</strong> at {top_sal['median']}")
    # C-level count
    takeaways.append(f"<strong>{summary['c_level_count']}</strong> C-Level/EVP roles this period")
    # Growth hires
    takeaways.append(f"<strong>{summary['growth_pct']}%</strong> of openings are growth hires (net-new roles)")
    takeaway_html = "<br>".join(f"&bull; {t}" for t in takeaways[:5])

    return f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>The Monday Brief — ExecSignals</title>
</head>
<body style="margin:0;padding:0;background:#f4f4f7;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f4f4f7;">
        <tr>
            <td align="center" style="padding:24px 16px;">
                <table role="presentation" width="640" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,0.08);">

                    <!-- Header -->
                    <tr>
                        <td style="background:#0C0F1A;padding:28px 32px;text-align:center;">
                            <div style="font-size:13px;font-weight:600;color:#D4A054;letter-spacing:2px;text-transform:uppercase;margin-bottom:6px;">ExecSignals</div>
                            <h1 style="color:#fff;margin:0;font-size:22px;font-weight:700;letter-spacing:-0.3px;">The Monday Brief</h1>
                            <p style="color:#94A3B8;margin:6px 0 0;font-size:13px;">{html.escape(date_range)} &middot; {summary['total']} VP+ Leads</p>
                        </td>
                    </tr>

                    <!-- Summary Stats -->
                    <tr>
                        <td style="padding:20px 32px 16px;">
                            <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f8f9fa;border-radius:8px;">
                                <tr>
                                    <td width="25%" align="center" style="padding:16px 4px;">
                                        <div style="font-size:24px;font-weight:700;color:#0C0F1A;">{summary['total']}</div>
                                        <div style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:0.5px;">VP+ Leads</div>
                                    </td>
                                    <td width="25%" align="center" style="padding:16px 4px;">
                                        <div style="font-size:24px;font-weight:700;color:#0C0F1A;">{summary['avg_salary']}</div>
                                        <div style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:0.5px;">Avg Salary</div>
                                    </td>
                                    <td width="25%" align="center" style="padding:16px 4px;">
                                        <div style="font-size:24px;font-weight:700;color:#0C0F1A;">{summary['c_level_count']}</div>
                                        <div style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:0.5px;">C-Level Roles</div>
                                    </td>
                                    <td width="25%" align="center" style="padding:16px 4px;">
                                        <div style="font-size:24px;font-weight:700;color:#06D6A0;">{summary['growth_pct']}%</div>
                                        <div style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:0.5px;">Growth Hires</div>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>

                    <!-- Seniority + Segment -->
                    <tr>
                        <td style="padding:0 32px 16px;">
                            <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
                                <tr>
                                    <td width="48%" valign="top" style="padding-right:8px;">
                                        <div style="background:#f8f9fa;border-radius:6px;padding:14px;">
                                            <div style="font-size:10px;font-weight:700;color:#888;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;">By Seniority</div>
                                            <div style="font-size:13px;color:#444;line-height:1.9;">{seniority_html}</div>
                                        </div>
                                    </td>
                                    <td width="48%" valign="top" style="padding-left:8px;">
                                        <div style="background:#f8f9fa;border-radius:6px;padding:14px;">
                                            <div style="font-size:10px;font-weight:700;color:#888;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;">By Segment</div>
                                            <div style="font-size:13px;color:#444;line-height:1.9;">{segment_html}</div>
                                        </div>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>

                    <!-- TOP 10 LEADS -->
                    <tr>
                        <td style="padding:8px 32px 0;">
                            <div style="border-bottom:2px solid #D4A054;padding-bottom:8px;margin-bottom:0;">
                                <span style="font-size:13px;font-weight:700;color:#0C0F1A;text-transform:uppercase;letter-spacing:1px;">Top 10 Leads This Week</span>
                                <span style="font-size:11px;color:#888;float:right;">Salary | Signals | Freshness</span>
                            </div>
                        </td>
                    </tr>

                    {top5_html}

                    <!-- Leads 6-10 compact -->
                    <tr>
                        <td style="padding:10px 32px 4px;">
                            <div style="font-size:10px;font-weight:700;color:#888;text-transform:uppercase;letter-spacing:1px;margin-bottom:6px;">Also This Week (#6-10)</div>
                        </td>
                    </tr>
                    <tr>
                        <td style="padding:0 32px;">
                            <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="font-size:12px;">
                                {compact_html}
                            </table>
                        </td>
                    </tr>

                    <!-- Attachment Note -->
                    <tr>
                        <td style="padding:20px 32px 16px;text-align:center;">
                            <div style="background:#FFF8EE;border:1px solid #F0DFC0;border-radius:6px;padding:14px;">
                                <p style="margin:0;color:#B8863E;font-size:13px;font-weight:600;">
                                    2 attachments included with this brief:
                                </p>
                                <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin-top:8px;">
                                    <tr>
                                        <td width="50%" style="padding:4px 8px 4px 0;">
                                            <div style="background:#fff;border:1px solid #eee;border-radius:4px;padding:8px 10px;font-size:12px;">
                                                <strong style="color:#0C0F1A;">{html.escape(xlsx_name)}</strong><br>
                                                <span style="color:#999;font-size:11px;">{summary['total']} leads &middot; Color-coded scores &middot; Filterable</span>
                                            </div>
                                        </td>
                                        <td width="50%" style="padding:4px 0 4px 8px;">
                                            <div style="background:#fff;border:1px solid #eee;border-radius:4px;padding:8px 10px;font-size:12px;">
                                                <strong style="color:#0C0F1A;">{html.escape(pdf_name)}</strong><br>
                                                <span style="color:#999;font-size:11px;">1-page summary &middot; Forward to clients &middot; Print-ready</span>
                                            </div>
                                        </td>
                                    </tr>
                                </table>
                            </div>
                        </td>
                    </tr>

                    <!-- MARKET INTELLIGENCE -->
                    <tr>
                        <td style="padding:8px 32px 0;">
                            <div style="background:#0C0F1A;border-radius:8px;padding:16px 20px;">
                                <span style="font-size:13px;font-weight:700;color:#D4A054;text-transform:uppercase;letter-spacing:1.5px;">Market Intelligence</span>
                                <span style="font-size:11px;color:#94A3B8;display:block;margin-top:2px;">Updated every Monday at 6 AM ET</span>
                            </div>
                        </td>
                    </tr>

                    <!-- Salary Benchmarks -->
                    <tr>
                        <td style="padding:16px 32px 0;">
                            <div style="font-size:11px;font-weight:700;color:#D4A054;text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;border-bottom:1px solid #eee;padding-bottom:6px;">Salary Benchmarks &mdash; VP+ Median Max by Function</div>
                            <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="font-size:12px;">
                                <tr style="color:#888;font-size:10px;text-transform:uppercase;letter-spacing:0.5px;">
                                    <td style="padding:4px 0;font-weight:600;">Role</td>
                                    <td align="center" style="padding:4px 0;font-weight:600;">P25</td>
                                    <td align="center" style="padding:4px 0;font-weight:600;">Median</td>
                                    <td align="center" style="padding:4px 0;font-weight:600;">P75</td>
                                    <td align="right" style="padding:4px 0;font-weight:600;">4-Wk Trend</td>
                                </tr>
                                {salary_table}
                            </table>
                        </td>
                    </tr>

                    <!-- Velocity + Companies -->
                    <tr>
                        <td style="padding:20px 32px 0;">
                            <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
                                <tr>
                                    <td width="48%" valign="top" style="padding-right:8px;">
                                        <div style="font-size:11px;font-weight:700;color:#D4A054;text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;border-bottom:1px solid #eee;padding-bottom:6px;">Hiring Velocity by Industry</div>
                                        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="font-size:12px;">
                                            {velocity_html}
                                        </table>
                                    </td>
                                    <td width="48%" valign="top" style="padding-left:8px;">
                                        <div style="font-size:11px;font-weight:700;color:#D4A054;text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;border-bottom:1px solid #eee;padding-bottom:6px;">Top Hiring Companies</div>
                                        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="font-size:12px;">
                                            {companies_html}
                                        </table>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>

                    <!-- Geo + Stage + Stack -->
                    <tr>
                        <td style="padding:20px 32px 0;">
                            <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
                                <tr>
                                    <td width="48%" valign="top" style="padding-right:8px;">
                                        <div style="font-size:11px;font-weight:700;color:#D4A054;text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;border-bottom:1px solid #eee;padding-bottom:6px;">VP+ Leads by Metro</div>
                                        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="font-size:12px;">
                                            {geo_html}
                                        </table>
                                    </td>
                                    <td width="48%" valign="top" style="padding-left:8px;">
                                        <div style="font-size:11px;font-weight:700;color:#D4A054;text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;border-bottom:1px solid #eee;padding-bottom:6px;">Key Takeaways</div>
                                        <div style="font-size:12px;color:#444;line-height:1.8;">{takeaway_html}</div>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>

                    <!-- Feedback CTA -->
                    <tr>
                        <td style="padding:24px 32px;text-align:center;">
                            <p style="color:#555;font-size:13px;margin:0 0 12px;">
                                How useful was this week's brief? Hit reply or:
                            </p>
                            <a href="mailto:hello@execsignals.com?subject=Monday%20Brief%20Feedback"
                               style="display:inline-block;background:#0C0F1A;color:#fff;padding:10px 28px;border-radius:6px;text-decoration:none;font-size:13px;font-weight:600;">
                                Share Feedback
                            </a>
                        </td>
                    </tr>

                    <!-- Footer -->
                    <tr>
                        <td style="background:#0C0F1A;padding:20px 32px;">
                            <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
                                <tr>
                                    <td>
                                        <div style="font-size:13px;font-weight:600;color:#D4A054;margin-bottom:4px;">ExecSignals</div>
                                        <div style="font-size:11px;color:#94A3B8;line-height:1.6;">
                                            The Monday Brief &middot; ExecSignals<br>
                                            Pariter Media Inc.
                                        </div>
                                    </td>
                                    <td align="right" valign="top">
                                        <a href="mailto:hello@execsignals.com?subject=Unsubscribe" style="color:#94A3B8;font-size:11px;text-decoration:none;">Unsubscribe</a>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>

                </table>
            </td>
        </tr>
    </table>
</body>
</html>"""


def generate_email_text(leads, analytics, summary, date_str, date_range, ref_date=None, file_date=None):
    """Generate plain text version of The Monday Brief."""
    now = ref_date or datetime.now()
    fd = file_date or datetime.now().strftime("%b%d")
    lines = []
    lines.append("=" * 60)
    lines.append("EXECSIGNALS — THE MONDAY BRIEF")
    lines.append(f"{date_range} | {summary['total']} VP+ Leads")
    lines.append("=" * 60)
    lines.append("")
    lines.append(f"  {summary['total']} VP+ Leads  |  Avg Salary: {summary['avg_salary']}  |  "
                 f"C-Level: {summary['c_level_count']}  |  Growth Hires: {summary['growth_pct']}%")
    lines.append("")

    # Seniority
    lines.append("BY SENIORITY:")
    for tier, count in summary["seniority"].items():
        lines.append(f"  {tier}: {count}")
    lines.append("")

    lines.append("-" * 60)
    lines.append("TOP 10 LEADS THIS WEEK")
    lines.append("-" * 60)

    now = ref_date or datetime.now()
    for i, lead in enumerate(leads[:10], 1):
        title = lead.get("title") or "Untitled"
        company = format_company_name(lead.get("company_name"))
        salary = format_salary(lead.get("annual_salary_min"), lead.get("annual_salary_max"))
        location = clean_location(lead)
        score = lead["score"]
        job_url = get_best_job_url(lead)
        company_url = lead.get("company_url") or ""
        fee = estimate_placement_fee(lead)
        repost_count = lead.get("repost_count", 0)

        filtered_lead = {**lead, "signals": filter_signals_for_role(lead)}
        hiring_sig = extract_hiring_signal(filtered_lead)
        team_sig = extract_team_structure(filtered_lead)

        repost_tag = f"  [REPOSTED {repost_count}x]" if repost_count > 1 else ""
        search_tag = "  [RETAINED SEARCH]" if lead.get("is_search_firm") else ""
        lines.append(f"\n  #{i}{repost_tag}{search_tag}")
        lines.append(f"  {title}")
        lines.append(f"  {company}  |  {location}")
        fee_str = f"  |  Est. Fee: {fee}" if fee else ""
        lines.append(f"  {salary}{fee_str}")

        signals_str = "  |  ".join(filter(None, [hiring_sig, team_sig]))
        if signals_str:
            lines.append(f"  Signals: {signals_str}")

        lines.append(f"  Note: {generate_signal_note(filtered_lead)}")
        lines.append(f"  Apply: {job_url}")
        if company_url:
            lines.append(f"  Company: {company_url}")

    lines.append("")
    lines.append("-" * 60)
    lines.append("SALARY BENCHMARKS")
    lines.append("-" * 60)
    for b in analytics["salary_benchmarks"]:
        trend = b["trend_display"]
        lines.append(f"  {b['role']:20s} P25: {b['p25']:>6s}  Med: {b['median']:>6s}  "
                     f"P75: {b['p75']:>6s}  Trend: {trend}")

    lines.append("")
    lines.append("-" * 60)
    lines.append("HIRING VELOCITY BY INDUSTRY")
    lines.append("-" * 60)
    for v in analytics["industry_velocity"]:
        lines.append(f"  {v['industry']:25s} {v['count']:>4d} openings  ({v['wow_display']} WoW)")

    lines.append("")
    lines.append("-" * 60)
    lines.append(f"Attachments: ExecSignals_{fd}.xlsx + "
                 f"MarketIntel_{fd}.html")
    lines.append("")
    lines.append("Reply to this email with feedback.")
    lines.append("")
    lines.append("-" * 60)
    lines.append("The Monday Brief by ExecSignals | Pariter Media Inc.")
    lines.append("Reply with 'unsubscribe' to stop receiving.")
    lines.append("-" * 60)

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 7: MAIN CLI
# ═══════════════════════════════════════════════════════════════════════════════


def main():
    parser = argparse.ArgumentParser(
        description="ExecSignals — The Monday Brief Generator"
    )
    parser.add_argument("--preview", action="store_true",
                        help="Generate all files locally (no sending)")
    parser.add_argument("--send", action="store_true",
                        help="Generate + send via Resend")
    parser.add_argument("--db", default=DEFAULT_DB,
                        help=f"Path to jobs.db (default: {DEFAULT_DB})")
    parser.add_argument("--days", type=int, default=7,
                        help="Days back to search for leads (default: 7)")
    parser.add_argument("--top", type=int, default=50,
                        help="Limit to top N leads (default: 50)")
    parser.add_argument("--output-dir", default="output",
                        help="Output directory (default: output)")
    parser.add_argument("--resend-key",
                        help="Resend API key (or set RESEND_API_KEY env var)")
    args = parser.parse_args()

    if not args.preview and not args.send:
        parser.print_help()
        return

    if not os.path.exists(args.db):
        print(f"Error: Database not found at {args.db}")
        sys.exit(1)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── 1. Connect + fetch leads ──
    print(f"Connecting to {args.db}...")
    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row

    # Use latest data date for date headers; actual date for filenames
    ref_date = _get_data_reference_date(conn)
    print(f"Latest data date: {ref_date.strftime('%Y-%m-%d')}")
    date_str = f"{(ref_date - timedelta(days=args.days)).strftime('%b %d')} \u2013 {ref_date.strftime('%b %d, %Y')}"
    date_range = f"{(ref_date - timedelta(days=args.days)).strftime('%b %d')} - {ref_date.strftime('%b %d, %Y')}"
    file_date = datetime.now().strftime("%b%d")

    print(f"Fetching VP+ leads from last {args.days} days...")
    leads = fetch_hot_leads(args.db, args.days, "vp")

    if not leads:
        print("No leads found. Try increasing --days.")
        conn.close()
        sys.exit(0)

    # Correct seniority misclassifications, then score
    for lead in leads:
        correct_seniority(lead)
        lead["score"] = score_lead(lead)
        apply_freshness_bonus(lead, ref_date)
        lead["is_search_firm"] = is_search_firm(lead)
    leads.sort(key=lambda x: x["score"], reverse=True)

    print(f"Found {len(leads)} raw leads (before dedup)")

    # Deduplicate: same title+company posted in multiple locations → keep best
    leads = deduplicate_leads(leads)
    print(f"After dedup: {len(leads)} unique leads")

    # Filter false positives (training programs, internships, etc.)
    before_fp = len(leads)
    leads = [l for l in leads if not is_false_positive(l)]
    if before_fp - len(leads) > 0:
        print(f"Filtered {before_fp - len(leads)} false positives (training/internship programs)")

    # Repost detection: flag roles that appear across multiple scrape dates
    repost_counts = compute_repost_counts(conn)
    for lead in leads:
        key = (
            (lead.get("company_name_normalized") or "").lower(),
            (lead.get("title") or "").lower(),
        )
        lead["repost_count"] = repost_counts.get(key, 0)
    reposted = sum(1 for l in leads if l["repost_count"] > 1)
    print(f"Reposted roles (appeared in 2+ scrapes): {reposted}")

    if args.top:
        leads = leads[:args.top]

    scores = [l["score"] for l in leads]
    print(f"Score range: {min(scores)} - {max(scores)} (avg: {sum(scores)/len(scores):.1f})")
    print()

    # ── 2. Compute analytics ──
    print("Computing market analytics...")
    analytics = compute_all_analytics(conn, lead_days=args.days)
    summary = compute_summary_stats(leads, analytics["geo_breakdown"])
    conn.close()

    print(f"  Salary benchmarks: {len(analytics['salary_benchmarks'])} roles")
    print(f"  Industry velocity: {len(analytics['industry_velocity'])} industries")
    print(f"  Top companies: {len(analytics['top_companies'])} companies")
    print(f"  Geo breakdown: {len(analytics['geo_breakdown'])} metros")
    print(f"  Stack trends: {len(analytics['stack_trends'])} tools")
    print()

    # ── 3. Generate outputs ──
    print("Generating deliverables...")

    # CSV
    csv_path = output_dir / "hot_leads.csv"
    generate_csv(leads, str(csv_path))
    print(f"  CSV:           {csv_path} ({len(leads)} rows)")

    # Excel
    xlsx_path = output_dir / f"ExecSignals_{file_date}.xlsx"
    generate_excel(leads, analytics, str(xlsx_path), date_str, ref_date=ref_date)
    print(f"  Excel:         {xlsx_path}")

    # PDF one-pager (HTML)
    pdf_path = output_dir / f"MarketIntel_{file_date}.html"
    pdf_html = generate_market_intel_html(analytics, summary, date_str)
    with open(pdf_path, "w", encoding="utf-8") as f:
        f.write(pdf_html)
    print(f"  Market Intel:  {pdf_path} (open in browser → Print → Save as PDF)")

    # Email HTML
    email_html_path = output_dir / f"MondayBrief_{file_date}.html"
    email_html_content = generate_email_html(leads, analytics, summary, date_str, date_range, ref_date=ref_date, file_date=file_date)
    with open(email_html_path, "w", encoding="utf-8") as f:
        f.write(email_html_content)
    print(f"  Email HTML:    {email_html_path}")

    # Email text
    email_txt_path = output_dir / f"MondayBrief_{file_date}.txt"
    email_txt_content = generate_email_text(leads, analytics, summary, date_str, date_range, ref_date=ref_date, file_date=file_date)
    with open(email_txt_path, "w", encoding="utf-8") as f:
        f.write(email_txt_content)
    print(f"  Email text:    {email_txt_path}")

    print()

    # ── 4. Send (if requested) ──
    if args.send:
        print("Sending via Resend...")
        try:
            import resend
            api_key = args.resend_key or os.environ.get("RESEND_API_KEY")
            if not api_key:
                print("Error: Resend API key required. Use --resend-key or set RESEND_API_KEY")
                sys.exit(1)
            resend.api_key = api_key

            # For now, just confirm — subscriber list to be added later
            print("  Resend integration ready. Add subscribers to send.")
            print("  (Subscriber management coming with Phase B)")
        except ImportError:
            print("Error: 'resend' package not installed. Run: pip install resend")
            sys.exit(1)

    # ── Summary ──
    print("=" * 60)
    print(f"The Monday Brief generated successfully!")
    print(f"  {len(leads)} leads scored | {len(analytics['salary_benchmarks'])} salary benchmarks")
    print(f"  {len(analytics['industry_velocity'])} industries | {len(analytics['geo_breakdown'])} metros")
    print(f"  Output: {output_dir}/")
    print("=" * 60)


if __name__ == "__main__":
    main()
