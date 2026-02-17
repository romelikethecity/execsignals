#!/usr/bin/env python3
"""
Generate ExecSignals weekly delivery mockup — Feb 17, 2026
Premium $297/mo VP+ hiring intel workbook for executive recruiters.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUTPUT = "/Users/rome/Documents/projects/products/hot-leads/mockups/ExecSignals_Feb17.xlsx"

# ── Brand palette ──────────────────────────────────────────────────────────
NAVY = "0C0F1A"
AMBER = "D4A054"
LIGHT_AMBER = "E2A84B"
BLUE = "5B8DEF"
DARK_TEXT = "1A1A1A"
WHITE = "FFFFFF"
LIGHT_BG = "F8F9FB"
BORDER_COLOR = "D0D5DD"
RED_FONT = "C0392B"
AMBER_FONT = "B7791F"
GRAY_FONT = "888888"
GREEN_FONT = "27763D"
SECTION_BG = "1A1F2E"

# ── Fonts ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Plus Jakarta Sans", bold=True, color=WHITE, size=11)
BODY_FONT = Font(name="Plus Jakarta Sans", color=DARK_TEXT, size=10)
BODY_FONT_BOLD = Font(name="Plus Jakarta Sans", color=DARK_TEXT, size=10, bold=True)
LINK_FONT = Font(name="Plus Jakarta Sans", color="1155CC", size=10, underline="single")
SECTION_FONT = Font(name="DM Serif Display", bold=True, color=AMBER, size=13)
SUBSECTION_FONT = Font(name="Plus Jakarta Sans", bold=True, color=WHITE, size=11)

# ── Fills ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SCORE_GOLD = PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid")
SCORE_BLUE = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
SCORE_LIGHT = PatternFill(start_color=LIGHT_AMBER, end_color=LIGHT_AMBER, fill_type="solid")
SCORE_GRAY = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
ALT_ROW = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
SECTION_FILL = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")

# ── Borders ────────────────────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

# ── Alignments ─────────────────────────────────────────────────────────────
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

# ============================================================================
#  MOCK DATA
# ============================================================================

LEADS = [
    {
        "rank": 1, "score": 57, "title": "Chief Revenue Officer",
        "company": "Databricks", "location": "San Francisco, CA",
        "sal_min": 320000, "sal_max": 450000, "seniority": "C-Level",
        "signals": "Growth Hire, Build Team, Reports to CEO, Series I Funded",
        "days": 1, "url": "https://www.indeed.com/viewjob?jk=abc001",
        "stage": "Late Stage", "employees": 7200, "pub_priv": "Private",
        "note": "Post-Series I expansion — net-new CRO role, team build mandate"
    },
    {
        "rank": 2, "score": 54, "title": "Chief Financial Officer",
        "company": "Stripe", "location": "San Francisco, CA",
        "sal_min": 350000, "sal_max": 500000, "seniority": "C-Level",
        "signals": "IPO Prep, Board Mandate, Reports to CEO",
        "days": 1, "url": "https://www.indeed.com/viewjob?jk=abc002",
        "stage": "Late Stage", "employees": 8000, "pub_priv": "Private",
        "note": "Pre-IPO CFO hire — board-driven, likely replacing interim"
    },
    {
        "rank": 3, "score": 52, "title": "SVP, Global Sales",
        "company": "Snowflake", "location": "Bozeman, MT",
        "sal_min": 290000, "sal_max": 420000, "seniority": "SVP",
        "signals": "Growth Hire, Build Team, Quota Carrier, 100+ Reports",
        "days": 2, "url": "https://www.indeed.com/viewjob?jk=abc003",
        "stage": "Enterprise", "employees": 6800, "pub_priv": "Public",
        "note": "Replacing departed SVP — urgent fill, $2B+ quota org"
    },
    {
        "rank": 4, "score": 49, "title": "VP Engineering",
        "company": "Anthropic", "location": "San Francisco, CA",
        "sal_min": 380000, "sal_max": 550000, "seniority": "VP",
        "signals": "Growth Hire, Build Team, Reports to CTO, AI/ML Focus",
        "days": 1, "url": "https://www.indeed.com/viewjob?jk=abc004",
        "stage": "Late Stage", "employees": 1500, "pub_priv": "Private",
        "note": "Scaling engineering org 3x — new VP layer, massive budget"
    },
    {
        "rank": 5, "score": 47, "title": "Chief Marketing Officer",
        "company": "HubSpot", "location": "Boston, MA",
        "sal_min": 300000, "sal_max": 430000, "seniority": "C-Level",
        "signals": "Replacement, Board Mandate, P&L Owner",
        "days": 2, "url": "https://www.indeed.com/viewjob?jk=abc005",
        "stage": "Enterprise", "employees": 7400, "pub_priv": "Public",
        "note": "CMO departure announced Q4 — board wants enterprise pivot leader"
    },
    {
        "rank": 6, "score": 45, "title": "VP Sales, Enterprise",
        "company": "Figma", "location": "New York, NY",
        "sal_min": 260000, "sal_max": 380000, "seniority": "VP",
        "signals": "Growth Hire, Enterprise Push, Build Team, Reports to CRO",
        "days": 1, "url": "https://www.indeed.com/viewjob?jk=abc006",
        "stage": "Late Stage", "employees": 1800, "pub_priv": "Private",
        "note": "First dedicated enterprise VP — massive upmarket motion"
    },
    {
        "rank": 7, "score": 44, "title": "SVP Operations",
        "company": "DoorDash", "location": "San Francisco, CA",
        "sal_min": 280000, "sal_max": 400000, "seniority": "SVP",
        "signals": "Replacement, Scale Ops, Reports to COO, P&L Owner",
        "days": 3, "url": "https://www.indeed.com/viewjob?jk=abc007",
        "stage": "Enterprise", "employees": 19000, "pub_priv": "Public",
        "note": "Reorg under new COO — consolidating ops under single SVP"
    },
    {
        "rank": 8, "score": 43, "title": "VP Product",
        "company": "Notion", "location": "San Francisco, CA",
        "sal_min": 270000, "sal_max": 390000, "seniority": "VP",
        "signals": "Growth Hire, Build Team, AI Product Focus, Reports to CEO",
        "days": 2, "url": "https://www.indeed.com/viewjob?jk=abc008",
        "stage": "Late Stage", "employees": 800, "pub_priv": "Private",
        "note": "AI product strategy lead — CEO direct report, new role"
    },
    {
        "rank": 9, "score": 42, "title": "Chief People Officer",
        "company": "Coinbase", "location": "Remote (HQ: San Francisco)",
        "sal_min": 300000, "sal_max": 425000, "seniority": "C-Level",
        "signals": "Replacement, Culture Reset, Board Mandate",
        "days": 3, "url": "https://www.indeed.com/viewjob?jk=abc009",
        "stage": "Enterprise", "employees": 3500, "pub_priv": "Public",
        "note": "Post-layoff people rebuild — board wants retention-focused leader"
    },
    {
        "rank": 10, "score": 41, "title": "VP Data & Analytics",
        "company": "Plaid", "location": "San Francisco, CA",
        "sal_min": 250000, "sal_max": 360000, "seniority": "VP",
        "signals": "Growth Hire, Build Team, AI/ML Focus, Reports to CTO",
        "days": 2, "url": "https://www.indeed.com/viewjob?jk=abc010",
        "stage": "Late Stage", "employees": 1200, "pub_priv": "Private",
        "note": "New data org — splitting from eng, building 20-person team"
    },
    {
        "rank": 11, "score": 39, "title": "VP Marketing",
        "company": "Canva", "location": "Austin, TX",
        "sal_min": 230000, "sal_max": 340000, "seniority": "VP",
        "signals": "Growth Hire, Enterprise Push, Reports to CMO",
        "days": 3, "url": "https://www.indeed.com/viewjob?jk=abc011",
        "stage": "Late Stage", "employees": 5000, "pub_priv": "Private",
        "note": "Enterprise marketing build-out — US market expansion focus"
    },
    {
        "rank": 12, "score": 38, "title": "SVP Customer Success",
        "company": "Salesforce", "location": "San Francisco, CA",
        "sal_min": 275000, "sal_max": 395000, "seniority": "SVP",
        "signals": "Replacement, Retention Focus, 200+ Reports",
        "days": 4, "url": "https://www.indeed.com/viewjob?jk=abc012",
        "stage": "Enterprise", "employees": 73000, "pub_priv": "Public",
        "note": "Post-restructuring CS consolidation — churn reduction mandate"
    },
    {
        "rank": 13, "score": 36, "title": "VP Finance",
        "company": "Airtable", "location": "San Francisco, CA",
        "sal_min": 240000, "sal_max": 350000, "seniority": "VP",
        "signals": "Growth Hire, IPO Prep, Reports to CFO",
        "days": 2, "url": "https://www.indeed.com/viewjob?jk=abc013",
        "stage": "Late Stage", "employees": 900, "pub_priv": "Private",
        "note": "IPO readiness hire — building out finance org and controls"
    },
    {
        "rank": 14, "score": 35, "title": "VP Sales, Mid-Market",
        "company": "Monday.com", "location": "New York, NY",
        "sal_min": 220000, "sal_max": 320000, "seniority": "VP",
        "signals": "Growth Hire, Build Team, Quota Carrier",
        "days": 3, "url": "https://www.indeed.com/viewjob?jk=abc014",
        "stage": "Enterprise", "employees": 2100, "pub_priv": "Public",
        "note": "Segmenting sales org — new mid-market VP to own $50M ARR target"
    },
    {
        "rank": 15, "score": 34, "title": "Chief Technology Officer",
        "company": "Ramp", "location": "New York, NY",
        "sal_min": 340000, "sal_max": 480000, "seniority": "C-Level",
        "signals": "Growth Hire, Build Team, Reports to CEO, Fintech",
        "days": 4, "url": "https://www.indeed.com/viewjob?jk=abc015",
        "stage": "Growth", "employees": 700, "pub_priv": "Private",
        "note": "First external CTO — founder stepping back from day-to-day eng"
    },
    {
        "rank": 16, "score": 33, "title": "VP Partnerships & Alliances",
        "company": "Datadog", "location": "New York, NY",
        "sal_min": 230000, "sal_max": 330000, "seniority": "VP",
        "signals": "Growth Hire, Channel Build, Reports to CRO",
        "days": 5, "url": "https://www.indeed.com/viewjob?jk=abc016",
        "stage": "Enterprise", "employees": 5500, "pub_priv": "Public",
        "note": "Building partner ecosystem — first dedicated partnerships VP"
    },
    {
        "rank": 17, "score": 31, "title": "VP People Operations",
        "company": "Scale AI", "location": "San Francisco, CA",
        "sal_min": 220000, "sal_max": 310000, "seniority": "VP",
        "signals": "Growth Hire, Build Team, Reports to CPO",
        "days": 4, "url": "https://www.indeed.com/viewjob?jk=abc017",
        "stage": "Late Stage", "employees": 1100, "pub_priv": "Private",
        "note": "Hypergrowth people ops — 500 hires planned in next 12 months"
    },
    {
        "rank": 18, "score": 30, "title": "SVP Revenue Operations",
        "company": "Gong", "location": "San Francisco, CA",
        "sal_min": 250000, "sal_max": 360000, "seniority": "SVP",
        "signals": "Replacement, Systems Overhaul, Reports to CRO",
        "days": 5, "url": "https://www.indeed.com/viewjob?jk=abc018",
        "stage": "Late Stage", "employees": 1300, "pub_priv": "Private",
        "note": "RevOps rebuild after CRO change — full stack overhaul"
    },
    {
        "rank": 19, "score": 29, "title": "VP Legal & Compliance",
        "company": "Rippling", "location": "San Francisco, CA",
        "sal_min": 260000, "sal_max": 375000, "seniority": "VP",
        "signals": "Growth Hire, Regulatory Prep, Reports to GC",
        "days": 6, "url": "https://www.indeed.com/viewjob?jk=abc019",
        "stage": "Late Stage", "employees": 2800, "pub_priv": "Private",
        "note": "International expansion compliance — EU/UK regulatory build"
    },
    {
        "rank": 20, "score": 28, "title": "VP Customer Experience",
        "company": "Toast", "location": "Boston, MA",
        "sal_min": 210000, "sal_max": 300000, "seniority": "VP",
        "signals": "Replacement, Retention Focus, Reports to COO",
        "days": 7, "url": "https://www.indeed.com/viewjob?jk=abc020",
        "stage": "Enterprise", "employees": 5200, "pub_priv": "Public",
        "note": "CX overhaul — NPS dropped 15pts, board flagged as priority"
    },
]

SALARY_BENCHMARKS = [
    ("VP Sales", "$245K", "$290K", "$380K", "+4.2%"),
    ("CFO", "$310K", "$375K", "$480K", "+2.8%"),
    ("VP Engineering", "$280K", "$340K", "$450K", "+6.1%"),
    ("VP Marketing", "$220K", "$270K", "$350K", "+1.9%"),
    ("VP Operations", "$215K", "$260K", "$340K", "-0.5%"),
    ("VP Product", "$250K", "$310K", "$400K", "+3.7%"),
    ("VP People/HR", "$200K", "$245K", "$320K", "+1.2%"),
]

INDUSTRY_VELOCITY = [
    ("Enterprise SaaS", 142, "+18%"),
    ("Fintech", 87, "+12%"),
    ("AI / ML", 134, "+31%"),
    ("Healthcare Tech", 56, "+7%"),
    ("E-Commerce / DTC", 38, "-4%"),
    ("Cybersecurity", 63, "+22%"),
]

TOP_COMPANIES = [
    ("Salesforce", 14), ("Google", 12), ("Databricks", 11),
    ("Snowflake", 9), ("Stripe", 8), ("HubSpot", 8),
    ("Meta", 7), ("Anthropic", 7), ("Datadog", 6), ("Figma", 5),
]

GEO_BREAKDOWN = [
    ("San Francisco / Bay Area", 87),
    ("New York Metro", 64),
    ("Austin, TX", 28),
    ("Boston, MA", 24),
    ("Seattle, WA", 22),
    ("Los Angeles, CA", 18),
    ("Denver / Boulder, CO", 14),
    ("Chicago, IL", 12),
]


def apply_cell(ws, row, col, value, font=None, fill=None, alignment=None,
               border=None, number_format=None, hyperlink=None):
    """Helper to set cell value and styles."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    if hyperlink:
        cell.hyperlink = hyperlink
    return cell


def build_top_leads(wb):
    """Sheet 1: Top Leads — scored, color-coded, filterable."""
    ws = wb.active
    ws.title = "Top Leads"
    ws.sheet_properties.tabColor = AMBER

    # Column definitions: (header, width)
    columns = [
        ("#", 5),
        ("Score", 8),
        ("Title", 28),
        ("Company", 18),
        ("Location", 26),
        ("Salary Min", 14),
        ("Salary Max", 14),
        ("Seniority", 12),
        ("Signals", 44),
        ("Days Posted", 13),
        ("Source", 14),
        ("Company Stage", 15),
        ("Employees", 12),
        ("Public/Private", 15),
        ("Signal Note", 54),
    ]

    # Set column widths
    for i, (_, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Header row
    for i, (header, _) in enumerate(columns, 1):
        apply_cell(ws, 1, i, header,
                   font=HEADER_FONT, fill=HEADER_FILL,
                   alignment=CENTER, border=THIN_BORDER)

    # Row height for header
    ws.row_dimensions[1].height = 32

    # Data rows
    for idx, lead in enumerate(LEADS):
        row = idx + 2
        ws.row_dimensions[row].height = 30

        # Alternate row fill
        row_fill = ALT_ROW if idx % 2 == 1 else None

        # Score-based styling
        score = lead["score"]
        if score >= 40:
            score_fill = SCORE_GOLD
            score_font = Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11)
        elif score >= 30:
            score_fill = SCORE_BLUE
            score_font = Font(name="Plus Jakarta Sans", bold=True, color=WHITE, size=11)
        elif score >= 20:
            score_fill = SCORE_LIGHT
            score_font = Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11)
        else:
            score_fill = SCORE_GRAY
            score_font = Font(name="Plus Jakarta Sans", color="555555", size=11)

        # Days Posted color
        days = lead["days"]
        if days <= 2:
            days_font = Font(name="Plus Jakarta Sans", bold=True, color=RED_FONT, size=10)
        elif days <= 4:
            days_font = Font(name="Plus Jakarta Sans", bold=True, color=AMBER_FONT, size=10)
        else:
            days_font = Font(name="Plus Jakarta Sans", color=GRAY_FONT, size=10)

        # Seniority color
        sen = lead["seniority"]
        if sen == "C-Level":
            sen_font = Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=10)
        elif sen == "SVP":
            sen_font = Font(name="Plus Jakarta Sans", bold=True, color="5B8DEF", size=10)
        else:
            sen_font = BODY_FONT

        # Build cell values: (value, col_alignment, font_override, fill_override)
        values = [
            (lead["rank"], CENTER, BODY_FONT, row_fill),
            (lead["score"], CENTER, score_font, score_fill),
            (lead["title"], LEFT, BODY_FONT_BOLD, row_fill),
            (lead["company"], LEFT, BODY_FONT, row_fill),
            (lead["location"], LEFT, BODY_FONT, row_fill),
            (lead["sal_min"], RIGHT, BODY_FONT, row_fill),
            (lead["sal_max"], RIGHT, BODY_FONT, row_fill),
            (lead["seniority"], CENTER, sen_font, row_fill),
            (lead["signals"], LEFT_WRAP, BODY_FONT, row_fill),
            (lead["days"], CENTER, days_font, row_fill),
            ("View Job", CENTER, LINK_FONT, row_fill),
            (lead["stage"], CENTER, BODY_FONT, row_fill),
            (lead["employees"], RIGHT, BODY_FONT, row_fill),
            (lead["pub_priv"], CENTER, BODY_FONT, row_fill),
            (lead["note"], LEFT_WRAP, BODY_FONT, row_fill),
        ]

        for col_idx, (val, align, font, fill) in enumerate(values, 1):
            cell = apply_cell(ws, row, col_idx, val,
                              font=font, fill=fill, alignment=align,
                              border=THIN_BORDER)
            # Salary formatting
            if col_idx in (6, 7):
                cell.number_format = '$#,##0'
            # Employee count formatting
            if col_idx == 13:
                cell.number_format = '#,##0'

        # Hyperlink for Source column (col 11)
        source_cell = ws.cell(row=row, column=11)
        source_cell.hyperlink = lead["url"]
        source_cell.value = "View Job"
        source_cell.font = LINK_FONT

    # Auto-filter on all columns
    ws.auto_filter.ref = f"A1:O{len(LEADS) + 1}"

    # Freeze top row
    ws.freeze_panes = "A2"

    ws.sheet_view.zoomScale = 100


def build_market_intel(wb):
    """Sheet 2: Market Intel — salary benchmarks, velocity, top companies, geo."""
    ws = wb.create_sheet("Market Intel")
    ws.sheet_properties.tabColor = "5B8DEF"

    # Column widths
    col_widths = {1: 4, 2: 30, 3: 14, 4: 14, 5: 14, 6: 16,
                  7: 4, 8: 30, 9: 14, 10: 14}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    current_row = 1

    # ── Title bar ──────────────────────────────────────────────────────────
    ws.merge_cells(f"B{current_row}:F{current_row}")
    title_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    apply_cell(ws, current_row, 2,
               "ExecSignals  \u2014  Market Intelligence Brief  |  Week of Feb 17, 2026",
               font=Font(name="DM Serif Display", bold=True, color=AMBER, size=14),
               fill=title_fill,
               alignment=Alignment(horizontal="left", vertical="center"))
    for c in range(3, 7):
        apply_cell(ws, current_row, c, None, fill=title_fill)
    ws.row_dimensions[current_row].height = 42
    current_row += 1

    # Subtitle line
    ws.merge_cells(f"B{current_row}:F{current_row}")
    apply_cell(ws, current_row, 2,
               "VP+ hiring intelligence sourced from 440K+ job postings  |  269 new VP+ roles this week",
               font=Font(name="Plus Jakarta Sans", italic=True, color=GRAY_FONT, size=9),
               alignment=LEFT)
    ws.row_dimensions[current_row].height = 20
    current_row += 2

    # ── SECTION 1: Salary Benchmarks ───────────────────────────────────────
    ws.merge_cells(f"B{current_row}:F{current_row}")
    apply_cell(ws, current_row, 2, "  SALARY BENCHMARKS \u2014 VP+ ROLES",
               font=Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=12),
               fill=SECTION_FILL,
               alignment=Alignment(horizontal="left", vertical="center"))
    for c in range(3, 7):
        apply_cell(ws, current_row, c, None, fill=SECTION_FILL)
    ws.row_dimensions[current_row].height = 32
    current_row += 1

    # Column headers
    bench_headers = ["Role", "P25", "Median", "P75", "4-Week Trend"]
    for i, h in enumerate(bench_headers):
        apply_cell(ws, current_row, i + 2, h,
                   font=HEADER_FONT, fill=HEADER_FILL,
                   alignment=CENTER, border=THIN_BORDER)
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    for idx, (role, p25, med, p75, trend) in enumerate(SALARY_BENCHMARKS):
        fill = ALT_ROW if idx % 2 == 1 else None

        if trend.startswith("+"):
            trend_font = Font(name="Plus Jakarta Sans", bold=True, color=GREEN_FONT, size=10)
            trend_display = f"\u25B2 {trend}"
        elif trend.startswith("-"):
            trend_font = Font(name="Plus Jakarta Sans", bold=True, color=RED_FONT, size=10)
            trend_display = f"\u25BC {trend}"
        else:
            trend_font = BODY_FONT
            trend_display = trend

        vals = [
            (role, LEFT, BODY_FONT_BOLD),
            (p25, RIGHT, BODY_FONT),
            (med, RIGHT, Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11)),
            (p75, RIGHT, BODY_FONT),
            (trend_display, CENTER, trend_font),
        ]
        for ci, (v, al, fnt) in enumerate(vals):
            apply_cell(ws, current_row, ci + 2, v,
                       font=fnt, fill=fill, alignment=al, border=THIN_BORDER)
        ws.row_dimensions[current_row].height = 26
        current_row += 1

    current_row += 2

    # ── SECTION 2: Hiring Velocity by Industry ─────────────────────────────
    ws.merge_cells(f"B{current_row}:F{current_row}")
    apply_cell(ws, current_row, 2, "  HIRING VELOCITY BY INDUSTRY",
               font=Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=12),
               fill=SECTION_FILL,
               alignment=Alignment(horizontal="left", vertical="center"))
    for c in range(3, 7):
        apply_cell(ws, current_row, c, None, fill=SECTION_FILL)
    ws.row_dimensions[current_row].height = 32
    current_row += 1

    vel_headers = ["Industry", "VP+ Openings This Week", "WoW Change", "", ""]
    for i, h in enumerate(vel_headers):
        apply_cell(ws, current_row, i + 2, h,
                   font=HEADER_FONT, fill=HEADER_FILL,
                   alignment=CENTER, border=THIN_BORDER)
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    for idx, (industry, count, wow) in enumerate(INDUSTRY_VELOCITY):
        fill = ALT_ROW if idx % 2 == 1 else None
        if wow.startswith("+") and int(wow.replace("+", "").replace("%", "")) >= 20:
            wow_font = Font(name="Plus Jakarta Sans", bold=True, color=GREEN_FONT, size=10)
            wow_display = f"\u25B2 {wow}"
        elif wow.startswith("+"):
            wow_font = Font(name="Plus Jakarta Sans", color=GREEN_FONT, size=10)
            wow_display = f"\u25B2 {wow}"
        else:
            wow_font = Font(name="Plus Jakarta Sans", bold=True, color=RED_FONT, size=10)
            wow_display = f"\u25BC {wow}"

        apply_cell(ws, current_row, 2, industry,
                   font=BODY_FONT_BOLD, fill=fill, alignment=LEFT, border=THIN_BORDER)
        apply_cell(ws, current_row, 3, count,
                   font=Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11),
                   fill=fill, alignment=CENTER, border=THIN_BORDER)
        apply_cell(ws, current_row, 4, wow_display,
                   font=wow_font, fill=fill, alignment=CENTER, border=THIN_BORDER)
        apply_cell(ws, current_row, 5, None, fill=fill, border=THIN_BORDER)
        apply_cell(ws, current_row, 6, None, fill=fill, border=THIN_BORDER)
        ws.row_dimensions[current_row].height = 26
        current_row += 1

    current_row += 2

    # ── SECTION 3: Top Companies + Geo Breakdown (side by side) ────────────
    section_start = current_row

    # Top Companies header
    ws.merge_cells(f"B{current_row}:D{current_row}")
    apply_cell(ws, current_row, 2, "  TOP HIRING COMPANIES",
               font=Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=12),
               fill=SECTION_FILL,
               alignment=Alignment(horizontal="left", vertical="center"))
    for c in [3, 4]:
        apply_cell(ws, current_row, c, None, fill=SECTION_FILL)

    # Geo header (right side)
    ws.merge_cells(f"F{current_row}:H{current_row}")
    apply_cell(ws, current_row, 6, "  GEO BREAKDOWN",
               font=Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=12),
               fill=SECTION_FILL,
               alignment=Alignment(horizontal="left", vertical="center"))
    for c in [7, 8]:
        apply_cell(ws, current_row, c, None, fill=SECTION_FILL)

    ws.column_dimensions["H"].width = 14

    ws.row_dimensions[current_row].height = 32
    current_row += 1

    # Sub-headers
    for i, h in enumerate(["Company", "VP+ Roles", ""]):
        apply_cell(ws, current_row, i + 2, h,
                   font=HEADER_FONT, fill=HEADER_FILL,
                   alignment=CENTER, border=THIN_BORDER)

    for i, h in enumerate(["Metro Area", "VP+ Roles", ""]):
        apply_cell(ws, current_row, i + 6, h,
                   font=HEADER_FONT, fill=HEADER_FILL,
                   alignment=CENTER, border=THIN_BORDER)

    ws.row_dimensions[current_row].height = 28
    current_row += 1

    max_rows = max(len(TOP_COMPANIES), len(GEO_BREAKDOWN))
    for idx in range(max_rows):
        fill = ALT_ROW if idx % 2 == 1 else None
        ws.row_dimensions[current_row].height = 26

        # Top companies (left)
        if idx < len(TOP_COMPANIES):
            company, count = TOP_COMPANIES[idx]
            apply_cell(ws, current_row, 2, company,
                       font=BODY_FONT_BOLD, fill=fill, alignment=LEFT, border=THIN_BORDER)
            apply_cell(ws, current_row, 3, count,
                       font=Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11),
                       fill=fill, alignment=CENTER, border=THIN_BORDER)
            apply_cell(ws, current_row, 4, None, fill=fill, border=THIN_BORDER)

        # Geo breakdown (right)
        if idx < len(GEO_BREAKDOWN):
            metro, gcount = GEO_BREAKDOWN[idx]
            apply_cell(ws, current_row, 6, metro,
                       font=BODY_FONT_BOLD, fill=fill, alignment=LEFT, border=THIN_BORDER)
            apply_cell(ws, current_row, 7, gcount,
                       font=Font(name="Plus Jakarta Sans", bold=True, color=DARK_TEXT, size=11),
                       fill=fill, alignment=CENTER, border=THIN_BORDER)
            apply_cell(ws, current_row, 8, None, fill=fill, border=THIN_BORDER)

        current_row += 1

    current_row += 2

    # ── SECTION 4: This Week's Key Takeaways ──────────────────────────────
    ws.merge_cells(f"B{current_row}:H{current_row}")
    apply_cell(ws, current_row, 2, "  THIS WEEK'S KEY TAKEAWAYS",
               font=Font(name="Plus Jakarta Sans", bold=True, color=AMBER, size=12),
               fill=SECTION_FILL,
               alignment=Alignment(horizontal="left", vertical="center"))
    for c in range(3, 9):
        apply_cell(ws, current_row, c, None, fill=SECTION_FILL)
    ws.row_dimensions[current_row].height = 32
    current_row += 1

    takeaways = [
        "\u25CF  AI/ML hiring surged 31% WoW — Anthropic, Scale AI, and Databricks driving bulk of VP+ demand",
        "\u25CF  4 C-Level roles posted in <48 hours (CRO, CFO, CMO, CPO) — highest same-week C-suite volume in 6 weeks",
        "\u25CF  VP Engineering comp hit new high: $340K median (+6.1%) — AI premium pulling up entire function",
        "\u25CF  E-Commerce/DTC is the only sector contracting (-4% WoW) — continued post-holiday pullback",
        "\u25CF  San Francisco dominates with 87 VP+ openings (32% of total) — NYC distant second at 64",
        "\u25CF  \"Build Team\" signal appeared in 11 of 20 top leads — companies investing in org growth, not just backfills",
    ]

    for tk in takeaways:
        ws.merge_cells(f"B{current_row}:H{current_row}")
        apply_cell(ws, current_row, 2, tk,
                   font=Font(name="Plus Jakarta Sans", color=DARK_TEXT, size=10),
                   alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
        for c in range(3, 9):
            apply_cell(ws, current_row, c, None)
        ws.row_dimensions[current_row].height = 24
        current_row += 1

    current_row += 2

    # ── Footer ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"B{current_row}:H{current_row}")
    apply_cell(ws, current_row, 2,
               "ExecSignals  |  The Monday Brief  |  execsignals.com  |  Data sourced from 440K+ job postings across 3 metros",
               font=Font(name="Plus Jakarta Sans", italic=True, color=GRAY_FONT, size=9),
               alignment=LEFT)
    current_row += 1
    ws.merge_cells(f"B{current_row}:H{current_row}")
    apply_cell(ws, current_row, 2,
               "Confidential \u2014 for subscriber use only. Do not redistribute.",
               font=Font(name="Plus Jakarta Sans", italic=True, color=RED_FONT, size=9),
               alignment=LEFT)

    # Freeze top row
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def main():
    wb = openpyxl.Workbook()
    build_top_leads(wb)
    build_market_intel(wb)
    wb.save(OUTPUT)
    print(f"Workbook saved: {OUTPUT}")
    print(f"  Sheet 1: 'Top Leads' — {len(LEADS)} scored leads, color-coded, filterable")
    print(f"  Sheet 2: 'Market Intel' — salary benchmarks, industry velocity, top companies, geo breakdown, key takeaways")


if __name__ == "__main__":
    main()
